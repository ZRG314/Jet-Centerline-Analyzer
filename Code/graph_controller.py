"""Controller for graph computations, rendering, and graph export actions."""

import csv
import os
import tkinter as tk
import zipfile
from datetime import datetime
from statistics import NormalDist
from tkinter import filedialog, messagebox, simpledialog, ttk
from xml.etree import ElementTree as ET

import cv2
import numpy as np
try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    load_workbook = None

from graph_math import build_fit_equation, compute_best_fit as compute_best_fit_profile
from graph_math import format_graph_value as format_graph_value_text
from plotting import resolve_axis_limits as resolve_plot_axis_limits


class GraphController:
    def __init__(self, app):
        self.app = app

    def has_imported_profile_data(self):
        data = getattr(self.app, "imported_profile_data", None)
        return isinstance(data, dict) and bool(data)

    def has_imported_distribution_data(self):
        data = getattr(self.app, "imported_distribution_data", None)
        return isinstance(data, dict) and bool(data)

    def _parse_float(self, value):
        text = str(value or "").strip()
        if not text:
            return None
        try:
            return float(text)
        except (TypeError, ValueError):
            return None

    def _parse_int(self, value):
        parsed = self._parse_float(value)
        if parsed is None:
            return None
        try:
            return int(round(parsed))
        except (TypeError, ValueError):
            return None

    def _normalized_header_map(self, header_row):
        header_map = {}
        for idx, value in enumerate(header_row):
            normalized = str(value or "").strip().lower()
            if normalized and normalized not in header_map:
                header_map[normalized] = idx
        return header_map

    def _find_first_header(self, rows):
        for row in rows:
            if any(str(cell).strip() for cell in row):
                return row
        return None

    def _find_first_nonempty_row_index(self, rows):
        for row_index, row in enumerate(rows):
            if any(str(cell).strip() for cell in row):
                return row_index
        return None

    def _find_header_index(self, rows, required_headers):
        required = {str(name).strip().lower() for name in required_headers}
        for row_index, row in enumerate(rows):
            header_map = self._normalized_header_map(row)
            if required.issubset(set(header_map)):
                return row_index, header_map
        return None, None

    def _extract_metadata_from_export_rows(self, rows, start_row, header_map):
        metadata = {}
        key_idx = header_map.get("metadata_key")
        if key_idx is None:
            key_idx = header_map.get("metadata")
        value_idx = header_map.get("metadata_value")
        if value_idx is None:
            value_idx = header_map.get("value")
        if key_idx is None or value_idx is None:
            return metadata
        for row in rows[start_row:]:
            if not any(str(cell).strip() for cell in row):
                continue
            if key_idx >= len(row):
                continue
            key = str(row[key_idx]).strip()
            if not key:
                continue
            value = str(row[value_idx]).strip() if value_idx < len(row) else ""
            metadata[key] = value
        return metadata

    def _find_export_header(self, rows):
        header_index, header_map = self._find_header_index(rows, {"metadata_key", "metadata_value"})
        if header_index is not None:
            return header_index, header_map
        return self._find_header_index(rows, {"metadata", "value"})

    def _find_profile_export_header(self, rows):
        for row_index, row in enumerate(rows):
            header_map = self._normalized_header_map(row)
            if "column_px" in header_map and "mean_y_graph_px" in header_map:
                return row_index, header_map

            has_generic_column = any(name.startswith("column_") for name in header_map)
            has_generic_mean = any(name.startswith("mean_y_") for name in header_map)
            if has_generic_column and has_generic_mean:
                return row_index, header_map
        return None, None

    def _extract_export_metadata(self, rows, header_index=None, header_map=None):
        if header_index is not None and header_map is not None:
            metadata = self._extract_metadata_from_export_rows(rows, header_index + 1, header_map)
            if metadata:
                return metadata

        header_index, header_map = self._find_export_header(rows)
        if header_index is not None:
            metadata = self._extract_metadata_from_export_rows(rows, header_index + 1, header_map)
            if metadata:
                return metadata

        for row_index, row in enumerate(rows):
            for col_index, cell in enumerate(row):
                if str(cell).strip().lower() == "export_type":
                    metadata = {}
                    blank_run = 0
                    for scan_row in rows[row_index:]:
                        key = str(scan_row[col_index]).strip() if col_index < len(scan_row) else ""
                        value = str(scan_row[col_index + 1]).strip() if (col_index + 1) < len(scan_row) else ""
                        if not key:
                            blank_run += 1
                            if blank_run >= 2:
                                break
                            continue
                        blank_run = 0
                        metadata[key] = value
                    return metadata
        return {}

    def _build_imported_profile_payload(self, x_values, mean_values, lower_values, upper_values,
                                        file_path, metadata=None, x_label=None, y_label=None):
        x_arr = np.asarray(x_values, dtype=np.float64)
        mean_arr = np.asarray(mean_values, dtype=np.float64)
        lower_arr = np.asarray(lower_values, dtype=np.float64)
        upper_arr = np.asarray(upper_values, dtype=np.float64)

        if x_arr.size == 0 or mean_arr.size == 0:
            raise ValueError("No usable graph rows were found in the CSV.")
        if not (x_arr.size == mean_arr.size == lower_arr.size == upper_arr.size):
            raise ValueError("Imported graph columns must all have the same number of rows.")

        order = np.argsort(x_arr)
        x_arr = x_arr[order]
        mean_arr = mean_arr[order]
        lower_arr = lower_arr[order]
        upper_arr = upper_arr[order]

        valid = np.isfinite(x_arr) & np.isfinite(mean_arr) & np.isfinite(lower_arr) & np.isfinite(upper_arr)
        if not np.any(valid):
            raise ValueError("The CSV did not contain any valid numeric graph rows.")

        metadata = dict(metadata or {})
        unit_label = (
            str(metadata.get("unit_label", "")).strip()
            or str(metadata.get("axis_unit", "")).strip()
            or "value"
        )
        stdev_multiplier = self._parse_float(metadata.get("stdev_multiplier"))
        if stdev_multiplier is None:
            stdev_multiplier = self.parse_graph_stdevs()

        title_text = os.path.basename(file_path)
        source_video = str(metadata.get("source_video", "")).strip()
        if source_video:
            title_text = f"Imported Profile: {os.path.basename(source_video)}"

        return {
            "source": "imported_csv",
            "source_path": file_path,
            "metadata": metadata,
            "title": title_text,
            "unit_label": unit_label,
            "stdevs": float(stdev_multiplier),
            "x_label": x_label or f"Horizontal Position ({unit_label})",
            "y_label": y_label or f"Vertical Position ({unit_label})",
            "x_all_values": x_arr,
            "mean_values": mean_arr,
            "lower_values": lower_arr,
            "upper_values": upper_arr,
            "valid": valid,
        }

    def _distribution_default_x_label(self, kind, unit_label):
        if kind == "Positions":
            return f"Vertical Position ({unit_label})"
        if kind == "Z-Scores":
            return "Z-Score"
        return f"Residual ({unit_label})"

    def _build_imported_distribution_payload(self, positions=None, residuals=None, z_scores=None, metadata=None):
        def clean_array(values):
            if values is None:
                return None
            arr = np.asarray(values, dtype=np.float64)
            arr = arr[np.isfinite(arr)]
            return arr if arr.size >= 3 else None

        positions = clean_array(positions)
        residuals = clean_array(residuals)
        z_scores = clean_array(z_scores)
        if positions is None and residuals is None and z_scores is None:
            return None

        metadata = dict(metadata or {})
        unit_label = str(metadata.get("position_unit", "")).strip() or str(metadata.get("unit_label", "")).strip() or "value"
        sample_count = 0
        for arr in (positions, residuals, z_scores):
            if arr is not None:
                sample_count = max(sample_count, int(arr.size))
        return {
            "source": "imported_csv",
            "combined_columns": True,
            "position_values": positions,
            "residual_values": residuals,
            "z_score_values": z_scores,
            "unit_label": unit_label,
            "column_count": self._parse_int(metadata.get("column_count")) or 0,
            "valid_column_count": self._parse_int(metadata.get("valid_column_count")) or 0,
            "sample_count": sample_count,
            "metadata": metadata,
        }

    def _parse_embedded_distribution_from_export_rows(self, rows, header_index, header_map, metadata):
        position_key = next((name for name in header_map if name.startswith("position_")), None)
        residual_key = next((name for name in header_map if name.startswith("residual_")), None)
        if position_key is None or residual_key is None or "z_score" not in header_map:
            return None

        positions = []
        residuals = []
        z_scores = []
        for row in rows[header_index + 1:]:
            if not any(str(cell).strip() for cell in row):
                continue
            position_val = self._parse_float(row[header_map[position_key]]) if header_map[position_key] < len(row) else None
            residual_val = self._parse_float(row[header_map[residual_key]]) if header_map[residual_key] < len(row) else None
            z_score_val = self._parse_float(row[header_map["z_score"]]) if header_map["z_score"] < len(row) else None
            if position_val is None or residual_val is None or z_score_val is None:
                continue
            positions.append(position_val)
            residuals.append(residual_val)
            z_scores.append(z_score_val)

        if not positions:
            return None

        distribution_metadata = dict(metadata or {})
        if not distribution_metadata.get("column_count") and distribution_metadata.get("histogram_column_count"):
            distribution_metadata["column_count"] = distribution_metadata.get("histogram_column_count")
        if not distribution_metadata.get("valid_column_count") and distribution_metadata.get("histogram_valid_column_count"):
            distribution_metadata["valid_column_count"] = distribution_metadata.get("histogram_valid_column_count")
        if not distribution_metadata.get("sample_count") and distribution_metadata.get("histogram_sample_count"):
            distribution_metadata["sample_count"] = distribution_metadata.get("histogram_sample_count")
        if not distribution_metadata.get("position_unit") and distribution_metadata.get("histogram_position_unit"):
            distribution_metadata["position_unit"] = distribution_metadata.get("histogram_position_unit")
        if not distribution_metadata.get("residual_unit") and distribution_metadata.get("histogram_residual_unit"):
            distribution_metadata["residual_unit"] = distribution_metadata.get("histogram_residual_unit")
        if not distribution_metadata.get("position_unit"):
            distribution_metadata["position_unit"] = position_key.split("position_", 1)[1].strip() or "value"
        if not distribution_metadata.get("residual_unit"):
            distribution_metadata["residual_unit"] = residual_key.split("residual_", 1)[1].strip() or distribution_metadata.get("position_unit", "value")
        return self._build_imported_distribution_payload(positions, residuals, z_scores, metadata=distribution_metadata)

    def _read_csv_rows(self, file_path):
        try:
            with open(file_path, "r", newline="", encoding="utf-8-sig") as handle:
                sample = handle.read(4096)
                handle.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
                rows = list(csv.reader(handle, dialect))
        except OSError as exc:
            raise OSError(f"Could not read delimited file:\n{exc}") from exc
        return rows

    def _xlsx_column_index_from_ref(self, cell_ref):
        letters = []
        for char in str(cell_ref or ""):
            if char.isalpha():
                letters.append(char.upper())
            elif letters:
                break
        index = 0
        for char in letters:
            index = (index * 26) + (ord(char) - ord("A") + 1)
        return max(0, index - 1)

    def _xlsx_join_text(self, element, namespace):
        if element is None:
            return ""
        parts = []
        for node in element.iter():
            if node.tag == f"{{{namespace}}}t" and node.text:
                parts.append(node.text)
        return "".join(parts)

    def _xlsx_parse_shared_strings(self, archive):
        shared_strings = []
        if "xl/sharedStrings.xml" not in archive.namelist():
            return shared_strings

        namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
        for item in root.findall(f".//{{{namespace}}}si"):
            shared_strings.append(self._xlsx_join_text(item, namespace))
        return shared_strings

    def _xlsx_parse_cell_value(self, cell, shared_strings, namespace):
        cell_type = cell.attrib.get("t", "")
        value_node = cell.find(f"{{{namespace}}}v")
        if cell_type == "inlineStr":
            return self._xlsx_join_text(cell.find(f"{{{namespace}}}is"), namespace)
        if value_node is None or value_node.text is None:
            return ""

        raw_value = value_node.text
        if cell_type == "s":
            try:
                index = int(raw_value)
            except (TypeError, ValueError):
                return ""
            return shared_strings[index] if 0 <= index < len(shared_strings) else ""
        if cell_type == "b":
            return "TRUE" if str(raw_value).strip() == "1" else "FALSE"
        if cell_type == "str":
            return raw_value

        numeric = self._parse_float(raw_value)
        if numeric is None:
            return raw_value
        if float(numeric).is_integer():
            return int(numeric)
        return numeric

    def _read_xlsx_rows_stdlib(self, file_path):
        main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        pkg_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"

        try:
            archive = zipfile.ZipFile(file_path, "r")
        except OSError as exc:
            raise OSError(f"Could not read Excel file:\n{exc}") from exc
        except zipfile.BadZipFile as exc:
            raise ValueError(f"Could not open Excel file:\n{exc}") from exc

        try:
            shared_strings = self._xlsx_parse_shared_strings(archive)

            workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
            rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            rel_map = {}
            for rel in rels_root.findall(f".//{{{pkg_rel_ns}}}Relationship"):
                rel_id = rel.attrib.get("Id")
                target = str(rel.attrib.get("Target", "")).replace("\\", "/")
                if target and not target.startswith("xl/"):
                    if target.startswith("/"):
                        target = target.lstrip("/")
                    else:
                        target = f"xl/{target}"
                if rel_id and target:
                    rel_map[rel_id] = target

            workbook_view = workbook_root.find(f".//{{{main_ns}}}workbookView")
            try:
                active_index = int(workbook_view.attrib.get("activeTab", "0")) if workbook_view is not None else 0
            except (TypeError, ValueError):
                active_index = 0

            sheets = []
            for sheet in workbook_root.findall(f".//{{{main_ns}}}sheet"):
                rel_id = sheet.attrib.get(f"{{{rel_ns}}}id")
                target = rel_map.get(rel_id)
                if target:
                    sheets.append(target)

            def worksheet_rows(path_in_archive):
                sheet_root = ET.fromstring(archive.read(path_in_archive))
                rows = []
                has_data = False
                for row_node in sheet_root.findall(f".//{{{main_ns}}}row"):
                    values_by_index = {}
                    max_index = -1
                    for cell in row_node.findall(f"{{{main_ns}}}c"):
                        col_index = self._xlsx_column_index_from_ref(cell.attrib.get("r", ""))
                        values_by_index[col_index] = self._xlsx_parse_cell_value(cell, shared_strings, main_ns)
                        max_index = max(max_index, col_index)
                    if max_index < 0:
                        normalized_row = []
                    else:
                        normalized_row = [values_by_index.get(idx, "") for idx in range(max_index + 1)]
                    if any(str(cell).strip() for cell in normalized_row):
                        has_data = True
                    rows.append(normalized_row)
                return rows, has_data

            if not sheets:
                return []

            active_index = max(0, min(active_index, len(sheets) - 1))
            active_rows, active_has_data = worksheet_rows(sheets[active_index])
            if active_has_data:
                return active_rows

            for idx, sheet_path in enumerate(sheets):
                if idx == active_index:
                    continue
                rows, has_data = worksheet_rows(sheet_path)
                if has_data:
                    return rows
            return active_rows
        except KeyError as exc:
            raise ValueError(f"Could not open Excel file:\nMissing workbook data: {exc}") from exc
        except ET.ParseError as exc:
            raise ValueError(f"Could not parse Excel file:\n{exc}") from exc
        finally:
            archive.close()

    def _read_xlsx_rows(self, file_path):
        if load_workbook is None:
            return self._read_xlsx_rows_stdlib(file_path)

        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        except OSError as exc:
            raise OSError(f"Could not read Excel file:\n{exc}") from exc
        except Exception as exc:
            return self._read_xlsx_rows_stdlib(file_path)

        try:
            def worksheet_rows(worksheet):
                rows = []
                has_data = False
                for row in worksheet.iter_rows(values_only=True):
                    normalized_row = [("" if cell is None else cell) for cell in row]
                    if any(str(cell).strip() for cell in normalized_row):
                        has_data = True
                    rows.append(normalized_row)
                return rows, has_data

            active_rows, active_has_data = worksheet_rows(workbook.active)
            if active_has_data:
                return active_rows

            for worksheet in workbook.worksheets:
                if worksheet == workbook.active:
                    continue
                rows, has_data = worksheet_rows(worksheet)
                if has_data:
                    return rows
            return active_rows
        finally:
            workbook.close()

    def _read_tabular_rows(self, file_path):
        ext = os.path.splitext(str(file_path or ""))[1].lower()
        if ext == ".xlsx":
            return self._read_xlsx_rows(file_path)
        return self._read_csv_rows(file_path)

    def _suggest_mapping_header_index(self, rows):
        export_header_index, _header_map = self._find_profile_export_header(rows)
        if export_header_index is not None:
            return export_header_index

        export_header_index, _header_map = self._find_export_header(rows)
        if export_header_index is not None:
            return export_header_index

        best_index = self._find_first_nonempty_row_index(rows)
        best_score = -1
        for row_index, row in enumerate(rows):
            score = sum(1 for cell in row if str(cell).strip())
            if score > best_score:
                best_score = score
                best_index = row_index
        return best_index

    def _build_column_option_labels(self, header_row):
        labels = []
        used = {}
        for idx, value in enumerate(header_row):
            base = str(value).strip() or f"Column {idx + 1}"
            count = used.get(base, 0) + 1
            used[base] = count
            label = base if count == 1 else f"{base} ({count})"
            labels.append(label)
        return labels

    def _guess_column_mapping_labels(self, rows, header_index, header_row):
        header_map = self._normalized_header_map(header_row)
        x_aliases = ("x", "x_value", "x_position", "column", "column_x", "column_px")
        mean_aliases = ("y", "mean", "mean_y", "mean_y_graph_px", "value")
        lower_aliases = ("lower", "lower_band", "lower_band_y", "lower_band_y_graph_px", "y_lower")
        upper_aliases = ("upper", "upper_band", "upper_band_y", "upper_band_y_graph_px", "y_upper")

        def first_index(*aliases, prefix=None):
            for alias in aliases:
                if alias in header_map:
                    return header_map[alias]
            if prefix is not None:
                for key, idx in header_map.items():
                    if key.startswith(prefix):
                        return idx
            return None

        guessed = {
            "x_index": first_index(*x_aliases, prefix="x_"),
            "mean_index": first_index(*mean_aliases, prefix="mean_y_"),
            "lower_index": first_index(*lower_aliases, prefix="lower_band_y_"),
            "upper_index": first_index(*upper_aliases, prefix="upper_band_y_"),
            "position_index": first_index(prefix="position_"),
            "residual_index": first_index(prefix="residual_"),
            "z_score_index": first_index("z_score"),
        }

        metadata = self._extract_export_metadata(rows, header_index, header_map)
        unit_label = str(metadata.get("unit_label", "")).strip()
        unit_suffix = unit_label.lower()
        if unit_suffix:
            unit_x_name = f"column_{unit_suffix}"
            unit_mean_name = f"mean_y_{unit_suffix}"
            unit_lower_name = f"lower_band_y_{unit_suffix}"
            unit_upper_name = f"upper_band_y_{unit_suffix}"
            if all(name in header_map for name in (unit_x_name, unit_mean_name)):
                guessed["x_index"] = header_map[unit_x_name]
                guessed["mean_index"] = header_map[unit_mean_name]
                guessed["lower_index"] = header_map[unit_lower_name] if unit_lower_name in header_map else None
                guessed["upper_index"] = header_map[unit_upper_name] if unit_upper_name in header_map else None

        return guessed

    def _prompt_for_csv_column_mapping(self, rows, file_path):
        header_index = self._suggest_mapping_header_index(rows)
        if header_index is None:
            raise ValueError("The imported file is empty.")

        header_row = rows[header_index]
        option_labels = self._build_column_option_labels(header_row)
        if not option_labels:
            raise ValueError("The imported file does not contain any columns to map.")

        label_to_index = {label: idx for idx, label in enumerate(option_labels)}
        guessed = self._guess_column_mapping_labels(rows, header_index, header_row)
        none_label = "(None)"

        popup = tk.Toplevel(self.app.root)
        popup.title("Map Imported Columns")
        popup.resizable(False, False)
        popup.transient(self.app.root)
        popup.grab_set()

        frame = tk.Frame(popup, padx=12, pady=10)
        frame.pack(fill="both", expand=True)
        frame.columnconfigure(1, weight=1)

        tk.Label(
            frame,
            text=(
                f"Choose which columns to import from {os.path.basename(file_path)}.\n"
                "X and Mean Y are required. Lower, Upper, and distribution columns are optional."
            ),
            justify="left",
            wraplength=420,
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 10))

        x_var = tk.StringVar(value=option_labels[guessed["x_index"]] if guessed["x_index"] is not None and guessed["x_index"] < len(option_labels) else option_labels[0])
        mean_default_index = guessed["mean_index"]
        if mean_default_index is None:
            mean_default_index = 1 if len(option_labels) > 1 else 0
        mean_var = tk.StringVar(value=option_labels[mean_default_index])
        lower_var = tk.StringVar(value=option_labels[guessed["lower_index"]] if guessed["lower_index"] is not None and guessed["lower_index"] < len(option_labels) else none_label)
        upper_var = tk.StringVar(value=option_labels[guessed["upper_index"]] if guessed["upper_index"] is not None and guessed["upper_index"] < len(option_labels) else none_label)
        position_var = tk.StringVar(value=option_labels[guessed["position_index"]] if guessed["position_index"] is not None and guessed["position_index"] < len(option_labels) else none_label)
        residual_var = tk.StringVar(value=option_labels[guessed["residual_index"]] if guessed["residual_index"] is not None and guessed["residual_index"] < len(option_labels) else none_label)
        z_score_var = tk.StringVar(value=option_labels[guessed["z_score_index"]] if guessed["z_score_index"] is not None and guessed["z_score_index"] < len(option_labels) else none_label)

        option_values = list(option_labels)
        optional_values = [none_label, *option_values]

        controls = [
            ("X Column", x_var, option_values),
            ("Mean Y Column", mean_var, option_values),
            ("Lower Band Column", lower_var, optional_values),
            ("Upper Band Column", upper_var, optional_values),
            ("Distribution Position Column", position_var, optional_values),
            ("Distribution Residual Column", residual_var, optional_values),
            ("Distribution Z-Score Column", z_score_var, optional_values),
        ]

        for row_offset, (label_text, var, values) in enumerate(controls, start=1):
            tk.Label(frame, text=label_text).grid(row=row_offset, column=0, sticky="w", pady=(0, 8))
            combo = ttk.Combobox(frame, textvariable=var, values=values, state="readonly", width=32)
            combo.grid(row=row_offset, column=1, sticky="we", padx=(10, 0), pady=(0, 8))

        status_var = tk.StringVar(value="")
        tk.Label(frame, textvariable=status_var, justify="left", fg="firebrick").grid(
            row=len(controls) + 1, column=0, columnspan=2, sticky="w", pady=(2, 0)
        )

        result = {"mapping": None}

        def on_import():
            x_label = x_var.get().strip()
            mean_label = mean_var.get().strip()
            lower_label = lower_var.get().strip()
            upper_label = upper_var.get().strip()
            position_label = position_var.get().strip()
            residual_label = residual_var.get().strip()
            z_score_label = z_score_var.get().strip()

            if x_label not in label_to_index or mean_label not in label_to_index:
                status_var.set("Choose valid columns for X and Mean Y.")
                return
            if x_label == mean_label:
                status_var.set("X and Mean Y must be different columns.")
                return

            result["mapping"] = {
                "header_index": header_index,
                "x_index": label_to_index[x_label],
                "mean_index": label_to_index[mean_label],
                "lower_index": None if lower_label == none_label else label_to_index.get(lower_label),
                "upper_index": None if upper_label == none_label else label_to_index.get(upper_label),
                "position_index": None if position_label == none_label else label_to_index.get(position_label),
                "residual_index": None if residual_label == none_label else label_to_index.get(residual_label),
                "z_score_index": None if z_score_label == none_label else label_to_index.get(z_score_label),
                "x_label": x_label,
                "y_label": mean_label,
            }
            popup.destroy()

        def on_cancel():
            popup.destroy()

        button_row = tk.Frame(frame)
        button_row.grid(row=len(controls) + 2, column=0, columnspan=2, sticky="e", pady=(12, 0))
        tk.Button(button_row, text="Cancel", command=on_cancel).pack(side="right")
        tk.Button(button_row, text="Import", command=on_import).pack(side="right", padx=(0, 6))

        popup.update_idletasks()
        width = popup.winfo_reqwidth()
        height = popup.winfo_reqheight()
        screen_w = popup.winfo_screenwidth()
        screen_h = popup.winfo_screenheight()
        x = max(0, (screen_w - width) // 2)
        y = max(0, (screen_h - height) // 2)
        popup.geometry(f"{width}x{height}+{x}+{y}")
        popup.bind("<Return>", lambda _event: on_import())
        popup.bind("<Escape>", lambda _event: on_cancel())
        popup.wait_window()
        return result["mapping"]

    def _parse_mapped_profile_csv(self, rows, file_path, mapping):
        if not mapping:
            return None

        header_index = int(mapping["header_index"])
        x_index = mapping["x_index"]
        mean_index = mapping["mean_index"]
        lower_index = mapping.get("lower_index")
        upper_index = mapping.get("upper_index")
        position_index = mapping.get("position_index")
        residual_index = mapping.get("residual_index")
        z_score_index = mapping.get("z_score_index")

        x_values = []
        mean_values = []
        lower_values = []
        upper_values = []
        positions = []
        residuals = []
        z_scores = []
        for row in rows[header_index + 1:]:
            if not any(str(cell).strip() for cell in row):
                continue

            x_val = self._parse_float(row[x_index]) if x_index < len(row) else None
            mean_val = self._parse_float(row[mean_index]) if mean_index < len(row) else None
            if x_val is None or mean_val is None:
                continue

            lower_val = self._parse_float(row[lower_index]) if lower_index is not None and lower_index < len(row) else mean_val
            upper_val = self._parse_float(row[upper_index]) if upper_index is not None and upper_index < len(row) else mean_val
            if lower_val is None:
                lower_val = mean_val
            if upper_val is None:
                upper_val = mean_val

            x_values.append(x_val)
            mean_values.append(mean_val)
            lower_values.append(lower_val)
            upper_values.append(upper_val)

            if position_index is not None and position_index < len(row):
                position_val = self._parse_float(row[position_index])
                if position_val is not None:
                    positions.append(position_val)
            if residual_index is not None and residual_index < len(row):
                residual_val = self._parse_float(row[residual_index])
                if residual_val is not None:
                    residuals.append(residual_val)
            if z_score_index is not None and z_score_index < len(row):
                z_score_val = self._parse_float(row[z_score_index])
                if z_score_val is not None:
                    z_scores.append(z_score_val)

        if not x_values:
            raise ValueError("No numeric data rows matched the selected column mapping.")

        profile_data = self._build_imported_profile_payload(
            x_values,
            mean_values,
            lower_values,
            upper_values,
            file_path,
            metadata={},
            x_label=str(mapping.get("x_label", "X")).strip() or "X",
            y_label=str(mapping.get("y_label", "Y")).strip() or "Y",
        )
        distribution_data = self._build_imported_distribution_payload(
            positions=positions if positions else None,
            residuals=residuals if residuals else None,
            z_scores=z_scores if z_scores else None,
            metadata={},
        )
        if distribution_data is not None:
            profile_data["distribution_data"] = distribution_data
        return profile_data

    def _parse_exported_profile_csv(self, rows, file_path):
        header_index, header_map = self._find_profile_export_header(rows)
        if header_index is None:
            return None

        metadata = self._extract_export_metadata(rows, header_index, header_map)
        unit_label = str(metadata.get("unit_label", "")).strip()
        unit_suffix = unit_label.lower()
        unit_x_name = f"column_{unit_suffix}" if unit_suffix else None
        unit_mean_name = f"mean_y_{unit_suffix}" if unit_suffix else None
        unit_lower_name = f"lower_band_y_{unit_suffix}" if unit_suffix else None
        unit_upper_name = f"upper_band_y_{unit_suffix}" if unit_suffix else None

        legacy_export = "column_px" in header_map and "mean_y_graph_px" in header_map
        generic_x_name = next((name for name in header_map if name.startswith("x_")), None)
        generic_mean_name = next((name for name in header_map if name.startswith("mean_y_")), None)
        generic_lower_name = next((name for name in header_map if name.startswith("lower_band_y_")), None)
        generic_upper_name = next((name for name in header_map if name.startswith("upper_band_y_")), None)

        use_unit_columns = all(
            name in header_map for name in (unit_x_name, unit_mean_name, unit_lower_name, unit_upper_name)
        ) if unit_suffix else False
        if not legacy_export and not (generic_x_name and generic_mean_name):
            return None

        x_values = []
        mean_values = []
        lower_values = []
        upper_values = []

        for row in rows[header_index + 1:]:
            if not any(str(cell).strip() for cell in row):
                continue

            if use_unit_columns:
                x_val = self._parse_float(row[header_map[unit_x_name]]) if header_map[unit_x_name] < len(row) else None
                mean_val = self._parse_float(row[header_map[unit_mean_name]]) if header_map[unit_mean_name] < len(row) else None
                lower_val = self._parse_float(row[header_map[unit_lower_name]]) if header_map[unit_lower_name] < len(row) else None
                upper_val = self._parse_float(row[header_map[unit_upper_name]]) if header_map[unit_upper_name] < len(row) else None
            elif legacy_export:
                x_val = self._parse_float(row[header_map["column_px"]]) if header_map["column_px"] < len(row) else None
                mean_val = self._parse_float(row[header_map["mean_y_graph_px"]]) if header_map["mean_y_graph_px"] < len(row) else None
                lower_val = self._parse_float(row[header_map["lower_band_y_graph_px"]]) if "lower_band_y_graph_px" in header_map and header_map["lower_band_y_graph_px"] < len(row) else mean_val
                upper_val = self._parse_float(row[header_map["upper_band_y_graph_px"]]) if "upper_band_y_graph_px" in header_map and header_map["upper_band_y_graph_px"] < len(row) else mean_val
            else:
                x_val = self._parse_float(row[header_map[generic_x_name]]) if header_map[generic_x_name] < len(row) else None
                mean_val = self._parse_float(row[header_map[generic_mean_name]]) if header_map[generic_mean_name] < len(row) else None
                lower_val = self._parse_float(row[header_map[generic_lower_name]]) if generic_lower_name is not None and header_map[generic_lower_name] < len(row) else mean_val
                upper_val = self._parse_float(row[header_map[generic_upper_name]]) if generic_upper_name is not None and header_map[generic_upper_name] < len(row) else mean_val

            if x_val is None or mean_val is None:
                continue
            if lower_val is None:
                lower_val = mean_val
            if upper_val is None:
                upper_val = mean_val

            x_values.append(x_val)
            mean_values.append(mean_val)
            lower_values.append(lower_val)
            upper_values.append(upper_val)

        if not x_values:
            return None

        if use_unit_columns and unit_label:
            axis_unit = unit_label
        elif legacy_export:
            axis_unit = "px"
        else:
            axis_unit = generic_x_name.split("x_", 1)[1].strip() if generic_x_name else (unit_label or "value")
        if not metadata.get("unit_label"):
            metadata["unit_label"] = axis_unit
        metadata["axis_unit"] = axis_unit
        profile_payload = self._build_imported_profile_payload(
            x_values,
            mean_values,
            lower_values,
            upper_values,
            file_path,
            metadata=metadata,
            x_label=f"Horizontal Position ({axis_unit})",
            y_label=f"Vertical Position ({axis_unit})",
        )
        distribution_payload = self._parse_embedded_distribution_from_export_rows(rows, header_index, header_map, metadata)
        if distribution_payload is not None:
            profile_payload["distribution_data"] = distribution_payload
        return profile_payload

    def _parse_simple_profile_csv(self, rows, file_path):
        header_row = self._find_first_header(rows)
        if header_row is None:
            return None

        header_map = self._normalized_header_map(header_row)
        x_aliases = ("x", "x_value", "x_position", "column", "column_x", "column_px")
        mean_aliases = ("y", "mean", "mean_y", "mean_y_graph_px", "value")
        lower_aliases = ("lower", "lower_band", "lower_band_y", "lower_band_y_graph_px", "y_lower")
        upper_aliases = ("upper", "upper_band", "upper_band_y", "upper_band_y_graph_px", "y_upper")

        def first_header(*aliases):
            for alias in aliases:
                if alias in header_map:
                    return alias
            return None

        x_key = first_header(*x_aliases)
        mean_key = first_header(*mean_aliases)
        lower_key = first_header(*lower_aliases)
        upper_key = first_header(*upper_aliases)
        if x_key is None:
            x_key = next((name for name in header_map if name.startswith("x_")), None)
        if mean_key is None:
            mean_key = next((name for name in header_map if name.startswith("mean_y_")), None)
        if lower_key is None:
            lower_key = next((name for name in header_map if name.startswith("lower_band_y_")), None)
        if upper_key is None:
            upper_key = next((name for name in header_map if name.startswith("upper_band_y_")), None)
        if x_key is None or mean_key is None:
            return None

        x_values = []
        mean_values = []
        lower_values = []
        upper_values = []
        header_index = rows.index(header_row)
        for row in rows[header_index + 1:]:
            if not any(str(cell).strip() for cell in row):
                continue
            x_val = self._parse_float(row[header_map[x_key]]) if header_map[x_key] < len(row) else None
            mean_val = self._parse_float(row[header_map[mean_key]]) if header_map[mean_key] < len(row) else None
            if x_val is None or mean_val is None:
                continue
            lower_val = self._parse_float(row[header_map[lower_key]]) if lower_key is not None and header_map[lower_key] < len(row) else mean_val
            upper_val = self._parse_float(row[header_map[upper_key]]) if upper_key is not None and header_map[upper_key] < len(row) else mean_val
            if lower_val is None:
                lower_val = mean_val
            if upper_val is None:
                upper_val = mean_val
            x_values.append(x_val)
            mean_values.append(mean_val)
            lower_values.append(lower_val)
            upper_values.append(upper_val)

        if not x_values:
            return None

        return self._build_imported_profile_payload(
            x_values,
            mean_values,
            lower_values,
            upper_values,
            file_path,
            metadata={},
            x_label=str(header_row[header_map[x_key]]).strip() or "X",
            y_label=str(header_row[header_map[mean_key]]).strip() or "Y",
        )

    def load_profile_csv_to_graph(self, file_path, force_mapping=False, rows=None):
        if rows is None:
            rows = self._read_tabular_rows(file_path)

        profile_data = None
        if not force_mapping:
            profile_data = self._parse_exported_profile_csv(rows, file_path)
            if profile_data is None:
                profile_data = self._parse_simple_profile_csv(rows, file_path)
        if profile_data is None:
            mapping = self._prompt_for_csv_column_mapping(rows, file_path)
            if mapping is None:
                return None
            profile_data = self._parse_mapped_profile_csv(rows, file_path, mapping)

        app = self.app
        app.imported_table_path = str(file_path or "")
        app.imported_table_rows = [list(row) for row in rows]
        app.imported_profile_data = profile_data
        app.imported_distribution_data = dict(profile_data.get("distribution_data", {})) if profile_data.get("distribution_data") else None
        app.graph_view_mode_var.set("Profile")
        app.final_centerline_samples = None
        app.graph_title_var.set("")
        x_label = profile_data.get("x_label")
        y_label = profile_data.get("y_label")
        if x_label:
            app.graph_x_axis_label.set(str(x_label).strip())
        if y_label:
            app.graph_y_axis_label.set(str(y_label).strip())
        stdevs = profile_data.get("stdevs")
        if stdevs is not None:
            app.graph_stdevs_var.set(self.format_graph_value(stdevs))

        unit_label = str(profile_data.get("unit_label", "")).strip()
        if unit_label:
            app.graph_units_label.configure(text=f"Units: {unit_label} (imported CSV)")

        self.redraw_graph()
        return profile_data

    def import_graph_csv(self, force_mapping=False):
        try:
            if force_mapping:
                file_path = str(getattr(self.app, "imported_table_path", "")).strip()
                rows = getattr(self.app, "imported_table_rows", None)
                if not file_path or not rows:
                    messagebox.showinfo("No file loaded", "Import a CSV or Excel file first, then use Map Imported Columns to remap it.")
                    return
                profile_data = self.load_profile_csv_to_graph(file_path, force_mapping=True, rows=rows)
            else:
                file_path = filedialog.askopenfilename(
                    title="Import Graph File",
                    filetypes=[
                        ("Graph Data Files", "*.csv *.xlsx"),
                        ("CSV file", "*.csv"),
                        ("Excel Workbook", "*.xlsx"),
                        ("All Files", "*.*"),
                    ],
                    initialdir=self.app.output_dir.get().strip() or os.path.dirname(__file__),
                )
                if not file_path:
                    return
                profile_data = self.load_profile_csv_to_graph(file_path, force_mapping=False)
            if profile_data is None:
                return
            messagebox.showinfo("File imported", f"Imported graph data from:\n{file_path}")
        except (OSError, ValueError) as exc:
            messagebox.showerror("Import failed", str(exc))

    def clear_imported_graph_data(self):
        app = self.app
        imported = getattr(app, "imported_profile_data", None) or {}
        imported_x_label = str(imported.get("x_label", "")).strip()
        imported_y_label = str(imported.get("y_label", "")).strip()

        app.imported_profile_data = None
        app.imported_distribution_data = None
        app.imported_table_path = ""
        app.imported_table_rows = None
        if imported_x_label and str(app.graph_x_axis_label.get()).strip() == imported_x_label:
            app.graph_x_axis_label.set(app.app_defaults["graph_x_axis_label"])
        if imported_y_label and str(app.graph_y_axis_label.get()).strip() == imported_y_label:
            app.graph_y_axis_label.set(app.app_defaults["graph_y_axis_label"])

        unit_label = str(app.graph_unit_label or "px").strip() or "px"
        if unit_label == "px":
            app.graph_units_label.configure(text="Units: px (uncalibrated)")
        else:
            app.graph_units_label.configure(
                text=f"Units: {unit_label} | Scale: {app.graph_unit_scale:.6g} {unit_label}/px"
            )
        self.redraw_graph()

    def graph_view_mode(self):
        return str(self.app.graph_view_mode_var.get()).strip()

    def resolve_graph_title(self, default_title):
        title_var = getattr(self.app, "graph_title_var", None)
        if title_var is None:
            return default_title
        custom = str(title_var.get()).strip()
        return custom if custom else default_title

    def edit_graph_title(self):
        title_var = getattr(self.app, "graph_title_var", None)
        if title_var is None:
            return
        current = str(title_var.get())
        new_value = simpledialog.askstring(
            "Edit Graph Title",
            "Enter graph title text (blank uses auto title).",
            initialvalue=current,
            parent=self.app.root,
        )
        if new_value is None:
            return
        title_var.set(str(new_value).strip())
        self.redraw_graph()

    def current_axis_label_text(self, axis_name):
        axis_name = str(axis_name).strip().upper()
        app = self.app
        mode = self.graph_view_mode()

        if mode == "Profile":
            if axis_name == "X":
                return app.graph_x_axis_label.get().strip() or self.profile_default_x_axis_label()
            return app.graph_y_axis_label.get().strip() or self.profile_default_y_axis_label()

        distribution_data = self.build_distribution_data()
        if distribution_data is None:
            return f"{axis_name} Axis"

        if mode == "Histogram":
            x_axis_text = distribution_data["default_x_label"]
            y_axis_text = distribution_data["default_y_label"]
        else:
            x_axis_text = "Theoretical Normal Quantile"
            y_axis_text = "Observed Quantile"

        return x_axis_text if axis_name == "X" else y_axis_text

    def _edit_axis_text(self, var, title, prompt):
        current_value = str(var.get())
        new_value = simpledialog.askstring(title, prompt, initialvalue=current_value, parent=self.app.root)
        if new_value is None:
            return
        var.set(str(new_value).strip())
        self.redraw_graph()

    def _split_axis_label_and_unit(self, label_text, fallback_label):
        text = str(label_text or "").strip()
        if text.endswith(")") and "(" in text:
            prefix, suffix = text.rsplit("(", 1)
            base = prefix.strip()
            unit = suffix[:-1].strip().lower()
            if base:
                if unit in ("mm", "pix", "px"):
                    unit = "pix" if unit in ("pix", "px") else "mm"
                    return base, unit
        return (text or fallback_label), "mm"

    def _axis_unit_from_var(self, var, fallback="mm"):
        text = str(var.get() if var is not None else "").strip()
        _base, unit = self._split_axis_label_and_unit(text, "")
        if unit in ("mm", "pix"):
            return unit
        return fallback

    def _clean_axis_bound_text(self, text, label):
        cleaned = str(text).strip()
        if not cleaned:
            return ""
        try:
            float(cleaned)
        except ValueError:
            messagebox.showerror("Invalid axis bound", f"{label} must be numeric or blank.")
            return None
        return cleaned

    def _resolve_auto_axis_limits(self, x_values, y_values, y_pad=1.0):
        return resolve_plot_axis_limits(
            x_values,
            y_values,
            x_user_min=None,
            x_user_max=None,
            y_user_min=None,
            y_user_max=None,
            y_pad=y_pad,
        )

    def current_graph_suggested_bounds(self):
        mode = self.graph_view_mode()

        if mode == "Profile":
            plot_data = self.build_plot_data()
            if plot_data is None:
                return None
            valid = plot_data["valid"]
            x_values = plot_data["x_all_values"][valid]
            y_values = np.concatenate((plot_data["upper_values"][valid], plot_data["lower_values"][valid]))
            return self._resolve_auto_axis_limits(x_values, y_values, y_pad=1.0)

        distribution_data = self.build_distribution_data()
        if distribution_data is None:
            return None

        if mode == "Histogram":
            if distribution_data.get("all_columns"):
                x_column_display = distribution_data["x_column_display"]
                edges = distribution_data["histogram_edges"]
                if x_column_display.size == 1:
                    column_span = 1.0
                    x_edges = np.array(
                        [
                            x_column_display[0] - 0.5 * column_span,
                            x_column_display[0] + 0.5 * column_span,
                        ],
                        dtype=np.float64,
                    )
                else:
                    midpoints = 0.5 * (x_column_display[:-1] + x_column_display[1:])
                    first_edge = x_column_display[0] - (midpoints[0] - x_column_display[0])
                    last_edge = x_column_display[-1] + (x_column_display[-1] - midpoints[-1])
                    x_edges = np.concatenate(([first_edge], midpoints, [last_edge]))
                return float(x_edges[0]), float(x_edges[-1]), float(edges[0]), float(edges[-1])

            edges = distribution_data["histogram_edges"]
            counts = distribution_data["histogram_counts"]
            limits = self._resolve_auto_axis_limits(edges, np.append(counts, 0), y_pad=1.0)
            if limits is None:
                return None
            x_min, x_max, _y_min, y_max = limits
            return x_min, x_max, 0.0, y_max

        theoretical = distribution_data["theoretical_quantiles"]
        observed = distribution_data["sorted_values"]
        combined = np.concatenate((theoretical, observed))
        return self._resolve_auto_axis_limits(combined, combined, y_pad=1.0)

    def x_axis_unit(self):
        fallback = "pix" if (self.app.graph_unit_label or "px").strip().lower() == "px" else "mm"
        return self._axis_unit_from_var(getattr(self.app, "graph_x_axis_label", None), fallback=fallback)

    def y_axis_unit(self):
        fallback = "pix" if (self.app.graph_unit_label or "px").strip().lower() == "px" else "mm"
        return self._axis_unit_from_var(getattr(self.app, "graph_y_axis_label", None), fallback=fallback)

    def _edit_axis_title_with_units(self, var, axis_name):
        axis_name = str(axis_name).strip().upper()
        current_value = str(var.get())
        visible_label = self.current_axis_label_text(axis_name)
        current_base, current_unit = self._split_axis_label_and_unit(current_value or visible_label, f"{axis_name} Axis")
        min_var = self.app.graph_x_min_var if axis_name == "X" else self.app.graph_y_min_var
        max_var = self.app.graph_x_max_var if axis_name == "X" else self.app.graph_y_max_var
        bound_min_var = tk.StringVar(value=str(min_var.get()).strip())
        bound_max_var = tk.StringVar(value=str(max_var.get()).strip())

        popup = tk.Toplevel(self.app.root)
        popup.title(f"Edit {axis_name} Axis Title")
        popup.resizable(False, False)
        popup.transient(self.app.root)
        popup.grab_set()

        title_var = tk.StringVar(value=current_base)
        unit_var = tk.StringVar(value=current_unit)

        frame = tk.Frame(popup, padx=12, pady=10)
        frame.pack(fill="both", expand=True)
        frame.columnconfigure(0, weight=1)

        tk.Label(frame, text="Title").grid(row=0, column=0, sticky="w")
        title_entry = tk.Entry(frame, textvariable=title_var, width=34)
        title_entry.grid(row=1, column=0, columnspan=2, sticky="we", pady=(2, 8))

        tk.Label(frame, text="Units").grid(row=2, column=0, sticky="w")
        unit_combo = ttk.Combobox(
            frame,
            textvariable=unit_var,
            values=["mm", "pix"],
            state="readonly",
            width=10,
        )
        unit_combo.grid(row=2, column=1, sticky="w", padx=(8, 0))

        tk.Label(frame, text=f"{axis_name} Axis Bounds").grid(row=3, column=0, columnspan=2, sticky="w", pady=(10, 0))
        tk.Label(frame, text="Min").grid(row=4, column=0, sticky="w", pady=(2, 0))
        min_entry = tk.Entry(frame, textvariable=bound_min_var, width=14)
        min_entry.grid(row=4, column=1, sticky="w", padx=(8, 0), pady=(2, 0))
        tk.Label(frame, text="Max").grid(row=5, column=0, sticky="w", pady=(6, 0))
        max_entry = tk.Entry(frame, textvariable=bound_max_var, width=14)
        max_entry.grid(row=5, column=1, sticky="w", padx=(8, 0), pady=(6, 0))

        suggested_bounds_var = tk.StringVar(value="Suggested min: n/a\nSuggested max: n/a")
        suggested_bounds_label = tk.Label(
            frame,
            textvariable=suggested_bounds_var,
            justify="left",
            fg="gray35",
        )
        suggested_bounds_label.grid(row=6, column=0, columnspan=2, sticky="w", pady=(6, 0))

        use_suggested_button = tk.Button(frame, text="Use Suggested Bounds")
        use_suggested_button.grid(row=7, column=0, columnspan=2, sticky="w", pady=(8, 0))

        button_row = tk.Frame(frame)
        button_row.grid(row=8, column=0, columnspan=2, sticky="e", pady=(12, 0))

        def refresh_suggested_bounds(*_args):
            selected_unit = unit_var.get().strip().lower()
            if selected_unit not in ("mm", "pix"):
                selected_unit = current_unit if current_unit in ("mm", "pix") else "mm"

            original_value = str(var.get())
            try:
                var.set(f"{current_base} ({selected_unit})")
                suggested_bounds = self.current_graph_suggested_bounds()
            finally:
                var.set(original_value)

            suggested_min = None
            suggested_max = None
            if suggested_bounds is not None:
                if axis_name == "X":
                    suggested_min, suggested_max = suggested_bounds[0], suggested_bounds[1]
                else:
                    suggested_min, suggested_max = suggested_bounds[2], suggested_bounds[3]

            suggested_min_text = self.format_graph_value(suggested_min) if suggested_min is not None else "n/a"
            suggested_max_text = self.format_graph_value(suggested_max) if suggested_max is not None else "n/a"
            suggested_bounds_var.set(f"Suggested min: {suggested_min_text}\nSuggested max: {suggested_max_text}")

            if suggested_min is not None and suggested_max is not None:
                use_suggested_button.configure(
                    state="normal",
                    command=lambda min_text=suggested_min_text, max_text=suggested_max_text: (
                        bound_min_var.set(min_text),
                        bound_max_var.set(max_text),
                    ),
                )
            else:
                use_suggested_button.configure(state="disabled", command=lambda: None)

        def on_save():
            cleaned_min = self._clean_axis_bound_text(bound_min_var.get(), f"{axis_name} axis minimum")
            if cleaned_min is None:
                return
            cleaned_max = self._clean_axis_bound_text(bound_max_var.get(), f"{axis_name} axis maximum")
            if cleaned_max is None:
                return
            if cleaned_min and cleaned_max and float(cleaned_max) <= float(cleaned_min):
                messagebox.showerror(
                    "Invalid axis bounds",
                    f"{axis_name} axis maximum must be greater than the minimum.",
                )
                return
            label_text = title_var.get().strip() or current_base
            selected_unit = unit_var.get().strip().lower()
            if selected_unit not in ("mm", "pix"):
                selected_unit = "mm"
            var.set(f"{label_text} ({selected_unit})")
            min_var.set(cleaned_min)
            max_var.set(cleaned_max)
            popup.destroy()
            self.redraw_graph()

        tk.Button(button_row, text="Cancel", command=popup.destroy).pack(side="right")
        tk.Button(button_row, text="Save", command=on_save).pack(side="right", padx=(0, 6))

        popup.update_idletasks()
        width = popup.winfo_reqwidth()
        height = popup.winfo_reqheight()
        screen_w = popup.winfo_screenwidth()
        screen_h = popup.winfo_screenheight()
        x = max(0, (screen_w - width) // 2)
        y = max(0, (screen_h - height) // 2)
        popup.geometry(f"{width}x{height}+{x}+{y}")

        popup.bind("<Return>", lambda _event: on_save())
        popup.bind("<Escape>", lambda _event: popup.destroy())
        unit_var.trace_add("write", refresh_suggested_bounds)
        refresh_suggested_bounds()
        title_entry.focus_set()
        popup.wait_window()

    def _edit_axis_limit(self, var, title):
        current_value = str(var.get())
        new_value = simpledialog.askstring(
            title,
            "Enter a number, or leave blank for auto bounds.",
            initialvalue=current_value,
            parent=self.app.root,
        )
        if new_value is None:
            return
        cleaned = str(new_value).strip()
        if cleaned:
            try:
                float(cleaned)
            except ValueError:
                messagebox.showerror("Invalid axis bound", "Axis bounds must be numeric or blank.")
                return
        var.set(cleaned)
        self.redraw_graph()

    def _bind_graph_edit_action(self, tag, callback):
        canvas = self.app.graph_canvas
        canvas.tag_bind(tag, "<Button-1>", lambda _event: callback())
        canvas.tag_bind(tag, "<Enter>", lambda _event: canvas.configure(cursor="hand2"))
        canvas.tag_bind(tag, "<Leave>", lambda _event: canvas.configure(cursor=""))

    def _clear_graph_edit_action(self, tag):
        canvas = self.app.graph_canvas
        for sequence in ("<Button-1>", "<Enter>", "<Leave>"):
            try:
                canvas.tag_unbind(tag, sequence)
            except tk.TclError:
                pass

    def _draw_editable_axis_controls(self, left, top, plot_w, plot_h):
        app = self.app
        canvas = app.graph_canvas
        axis_text_color = "black"

        x_label_text = self.current_axis_label_text("X")
        y_label_text = self.current_axis_label_text("Y")

        canvas.create_text(
            canvas.winfo_width() // 2,
            canvas.winfo_height() - 22,
            text=x_label_text,
            fill=axis_text_color,
            font=("TkDefaultFont", 10),
            tags=("graph_edit_x_label",),
        )
        canvas.create_text(
            28,
            canvas.winfo_height() // 2,
            text=y_label_text,
            fill=axis_text_color,
            angle=90,
            font=("TkDefaultFont", 10),
            tags=("graph_edit_y_label",),
        )

        self._bind_graph_edit_action(
            "graph_edit_x_label",
            lambda: self._edit_axis_title_with_units(app.graph_x_axis_label, "X"),
        )
        self._bind_graph_edit_action(
            "graph_edit_y_label",
            lambda: self._edit_axis_title_with_units(app.graph_y_axis_label, "Y"),
        )
        self._clear_graph_edit_action("graph_edit_x_min_tick")
        self._clear_graph_edit_action("graph_edit_x_max_tick")
        self._clear_graph_edit_action("graph_edit_y_min_tick")
        self._clear_graph_edit_action("graph_edit_y_max_tick")

    def histogram_scope(self):
        scope_var = getattr(self.app, "graph_histogram_scope_var", None)
        return str(scope_var.get()).strip() if scope_var is not None else "All Columns"

    def is_single_column_mode(self):
        mode = self.graph_view_mode()
        if mode == "Histogram" and self.histogram_scope() == "Selected Column":
            return True
        return False

    def profile_value_mode(self):
        return "Pixel Values" if self.y_axis_unit() == "pix" else "Actual Values"

    def profile_x_value_mode(self):
        return "Pixel Values" if self.x_axis_unit() == "pix" else "Actual Values"

    def column_input_mode(self):
        return "Pixel Values" if self.x_axis_unit() == "pix" else "Actual Values"

    def using_profile_pixel_values(self):
        return self.profile_value_mode() == "Pixel Values"

    def using_profile_x_pixel_values(self):
        return self.profile_x_value_mode() == "Pixel Values"

    def profile_x_value(self, x_px):
        if self.using_profile_x_pixel_values():
            return float(x_px)
        return float(self.x_position_to_graph_units(x_px))

    def profile_y_value(self, y_px):
        if self.using_profile_pixel_values():
            return float(self.y_position_to_graph_pixels(y_px))
        return float(self.y_position_to_graph_units(y_px))

    def profile_y_delta_value(self, dy_px):
        if self.using_profile_pixel_values():
            return float(dy_px)
        return float(self.y_delta_to_graph_units(dy_px))

    def profile_axis_unit_label(self):
        return "pix" if self.using_profile_pixel_values() else "mm"

    def profile_x_axis_unit_label(self):
        return "pix" if self.using_profile_x_pixel_values() else "mm"

    def profile_default_x_axis_label(self):
        return f"Horizontal Position ({self.profile_x_axis_unit_label()})"

    def profile_default_y_axis_label(self):
        return f"Vertical Position ({self.profile_axis_unit_label()})"

    def column_index_to_value(self, column_index):
        x_px = float(self.profile_index_to_x_px(column_index))
        if self.column_input_mode() == "Actual Values":
            return float(self.x_position_to_graph_units(x_px))
        return x_px

    def column_value_to_index(self, value, column_count):
        if column_count <= 0:
            return 0
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            numeric = 0.0
        x_px = numeric
        if self.column_input_mode() == "Actual Values":
            scale = float(self.app.graph_unit_scale) if float(self.app.graph_unit_scale) != 0 else 1.0
            x_px = numeric / scale
        index = int(round(x_px / self.pixels_per_column()))
        return max(0, min(int(column_count - 1), index))

    def column_input_suffix(self):
        return "pix" if self.column_input_mode() == "Pixel Values" else "mm"

    def format_column_input_bounds(self, column_count):
        if column_count <= 0:
            return "Input range: run analysis to populate bounds."
        low = self.column_index_to_value(0)
        high = self.column_index_to_value(column_count - 1)
        suffix = self.column_input_suffix()
        return f"Input range: {self.format_graph_value(low)} to {self.format_graph_value(high)} {suffix}"

    def refresh_distribution_column_controls(self):
        app = self.app
        mode = self.graph_view_mode()
        using_imported_distribution = self.has_imported_distribution_data()
        histogram_scope_combo = getattr(app, "graph_histogram_scope_combo", None)
        if mode == "Q-Q Plot" and getattr(app, "graph_histogram_scope_var", None) is not None:
            if str(app.graph_histogram_scope_var.get()).strip() != "All Columns (Combined)":
                app.graph_histogram_scope_var.set("All Columns (Combined)")
        elif using_imported_distribution and mode == "Histogram" and getattr(app, "graph_histogram_scope_var", None) is not None:
            if str(app.graph_histogram_scope_var.get()).strip() != "All Columns (Combined)":
                app.graph_histogram_scope_var.set("All Columns (Combined)")

        profile_mode = mode == "Profile"
        qq_mode = mode == "Q-Q Plot"

        def set_entry_enabled(widget, enabled):
            if widget is not None:
                widget.configure(state=("normal" if enabled else "disabled"))

        def set_combo_enabled(widget, enabled):
            if widget is not None:
                if isinstance(widget, ttk.Menubutton):
                    widget.configure(state=("normal" if enabled else "disabled"))
                else:
                    widget.configure(state=("readonly" if enabled else "disabled"))

        def set_label_enabled(widget, enabled, disabled_fg="gray55", enabled_fg="black"):
            if widget is not None:
                try:
                    widget.configure(text_color=(enabled_fg if enabled else disabled_fg))
                except Exception:
                    widget.configure(fg=(enabled_fg if enabled else disabled_fg))

        def set_header_enabled(header_row, enabled):
            if header_row is None:
                return
            for child in header_row.winfo_children():
                try:
                    child.configure(text_color=("black" if enabled else "gray55"))
                except Exception:
                    if isinstance(child, tk.Label):
                        child.configure(fg=("black" if enabled else "gray55"))

        distribution_controls_enabled = not profile_mode
        histogram_scope_enabled = not profile_mode

        distribution_kind_disabled_options = set()
        imported_distribution = getattr(app, "imported_distribution_data", None) or {}
        if hasattr(app, "configure_distribution_kind_menu"):
            if using_imported_distribution and not profile_mode:
                if imported_distribution.get("residual_values") is None:
                    distribution_kind_disabled_options.add("Residuals")
                if imported_distribution.get("position_values") is None:
                    distribution_kind_disabled_options.add("Positions")
                if imported_distribution.get("z_score_values") is None:
                    distribution_kind_disabled_options.add("Z-Scores")
            app.configure_distribution_kind_menu(disabled_options=distribution_kind_disabled_options)

        histogram_scope_disabled_options = set()
        if qq_mode or (using_imported_distribution and not profile_mode):
            histogram_scope_disabled_options.update({"All Columns", "Selected Column"})
        if hasattr(app, "configure_histogram_scope_menu"):
            app.configure_histogram_scope_menu(disabled_options=histogram_scope_disabled_options)

        set_combo_enabled(getattr(app, "graph_distribution_kind_combo", None), distribution_controls_enabled)
        set_header_enabled(getattr(app, "graph_distribution_kind_header_row", None), True)

        set_combo_enabled(getattr(app, "graph_histogram_scope_combo", None), histogram_scope_enabled)
        set_header_enabled(getattr(app, "graph_histogram_scope_header_row", None), histogram_scope_enabled)

        bins_enabled = not (profile_mode or qq_mode)
        set_entry_enabled(getattr(app, "graph_distribution_bins_entry", None), bins_enabled)
        set_label_enabled(getattr(app, "graph_distribution_bins_label", None), bins_enabled)

        if hasattr(app, "graph_distribution_column_label"):
            app.graph_distribution_column_label.configure(text=f"Selected Column ({self.column_input_suffix()})")

        single_column_mode = self.is_single_column_mode() and not using_imported_distribution
        column_enabled = distribution_controls_enabled and single_column_mode
        if hasattr(app, "graph_distribution_column_entry"):
            app.graph_distribution_column_entry.configure(state="normal" if column_enabled else "disabled")
        set_label_enabled(getattr(app, "graph_distribution_column_label", None), column_enabled)
        if hasattr(app, "graph_distribution_column_bounds_label"):
            app.graph_distribution_column_bounds_label.configure(fg=("gray35" if column_enabled else "gray55"))

        bounds_text = "Input range: run analysis to populate bounds."
        if using_imported_distribution and not profile_mode:
            bounds_text = "Imported CSV includes combined distribution data only."
        elif column_enabled:
            samples = getattr(app, "final_centerline_samples", None)
            if samples is not None:
                arr = np.asarray(samples, dtype=np.float64)
                if arr.ndim == 2 and arr.shape[1] > 0:
                    bounds_text = self.format_column_input_bounds(arr.shape[1])
        elif profile_mode:
            bounds_text = "Selected Column is available only for Histogram and Q-Q views."
        elif qq_mode:
            bounds_text = "Q-Q Plot always uses all columns combined."
        else:
            bounds_text = "Selected Column is active only for Single Column Histogram."
        if hasattr(app, "graph_distribution_column_bounds_var"):
            app.graph_distribution_column_bounds_var.set(bounds_text)

    def parse_graph_stdevs(self):
        app = self.app
        try:
            value = float(str(app.graph_stdevs_var.get()).strip())
            return max(0.0, value)
        except (TypeError, ValueError):
            return 0.0

    def parse_graph_fit_degree(self):
        app = self.app
        try:
            value = int(str(app.graph_fit_degree_var.get()).strip())
            return max(1, min(6, value))
        except (TypeError, ValueError):
            return 2

    def pixels_per_column(self):
        app = self.app
        value = app.safe_int(app.pixel_entry.get())
        if value is None or value <= 0:
            return 1.0
        return float(value)

    def profile_index_to_x_px(self, x_index):
        return float(x_index) * self.pixels_per_column()

    def to_graph_units(self, px_value):
        app = self.app
        return px_value * app.graph_unit_scale

    def x_position_to_graph_units(self, x_px):
        return self.to_graph_units(x_px)

    def get_profile_height_px(self):
        app = self.app
        if app.last_analysis_frame is not None:
            return int(app.last_analysis_frame.shape[0])
        height = int(app.crop_bottom - app.crop_top)
        if height > 0:
            return height
        return None

    def nozzle_origin_y_in_profile_px(self):
        app = self.app
        if app.nozzle_origin_img is None:
            return None
        nozzle_y = float(app.nozzle_origin_img[1])
        if app.crop_bottom > app.crop_top:
            return nozzle_y - float(app.crop_top)
        return nozzle_y

    def y_position_to_graph_units(self, y_px):
        app = self.app
        frame_height = self.get_profile_height_px()
        if frame_height is None or frame_height <= 1:
            return self.to_graph_units(y_px)
        
        nozzle_y = self.nozzle_origin_y_in_profile_px()
        if nozzle_y is not None:
            # With nozzle origin: measure absolute distance from nozzle
            return self.to_graph_units(abs(y_px - nozzle_y))
        
        # Without nozzle origin: flip y-axis so top of image = top of graph
        return self.to_graph_units((frame_height - 1) - y_px)

    def y_delta_to_graph_units(self, dy_px):
        return self.to_graph_units(dy_px)

    def y_position_to_graph_pixels(self, y_px):
        frame_height = self.get_profile_height_px()
        if frame_height is None or frame_height <= 1:
            return float(y_px)

        nozzle_y = self.nozzle_origin_y_in_profile_px()
        if nozzle_y is not None:
            return float(abs(y_px - nozzle_y))

        return float((frame_height - 1) - y_px)

    def parse_optional_float_var(self, var):
        text = str(var.get()).strip()
        if not text:
            return None
        try:
            return float(text)
        except (TypeError, ValueError):
            return None

    def resolve_axis_limits(self, x_values, y_values, y_pad=1.0):
        app = self.app
        x_user_min = self.parse_optional_float_var(app.graph_x_min_var)
        x_user_max = self.parse_optional_float_var(app.graph_x_max_var)
        y_user_min = self.parse_optional_float_var(app.graph_y_min_var)
        y_user_max = self.parse_optional_float_var(app.graph_y_max_var)
        return resolve_plot_axis_limits(
            x_values,
            y_values,
            x_user_min=x_user_min,
            x_user_max=x_user_max,
            y_user_min=y_user_min,
            y_user_max=y_user_max,
            y_pad=y_pad,
        )

    def format_graph_value(self, value):
        return format_graph_value_text(value)

    def compute_best_fit(self, mean, valid):
        return compute_best_fit_profile(mean, valid, self.parse_graph_fit_degree())

    def compute_imported_best_fit(self, x_values, mean_values, valid):
        x_values = np.asarray(x_values, dtype=np.float64)
        mean_values = np.asarray(mean_values, dtype=np.float64)
        valid = np.asarray(valid, dtype=bool)
        x_valid = x_values[valid]
        y_valid = mean_values[valid]
        if x_valid.size < 2:
            return None

        requested_degree = self.parse_graph_fit_degree()
        degree = max(1, min(6, int(requested_degree)))
        degree = min(degree, int(x_valid.size - 1))

        coeffs = np.polyfit(x_valid, y_valid, degree)
        poly = np.poly1d(coeffs)
        y_fit = poly(x_values)
        y_pred = poly(x_valid)

        ss_res = float(np.sum((y_valid - y_pred) ** 2))
        ss_tot = float(np.sum((y_valid - np.mean(y_valid)) ** 2))
        r2 = 1.0 - (ss_res / ss_tot) if ss_tot > 0 else 1.0
        equation = f"Best fit (degree {degree}): {build_fit_equation(coeffs)}   R^2={r2:.4f}"
        return {
            "degree": degree,
            "equation": equation,
            "y_fit_values": y_fit,
        }

    def compute_plot_fit(self, plot_data):
        source = str(plot_data.get("source", "native")).strip()
        if source == "imported_csv":
            return self.compute_imported_best_fit(
                plot_data["x_all_values"],
                plot_data["mean_values"],
                plot_data["valid"],
            )
        return self.compute_best_fit(plot_data["mean"], plot_data["valid"])

    def parse_distribution_bins(self):
        app = self.app
        try:
            value = int(str(app.graph_distribution_bins_var.get()).strip())
            return max(3, min(200, value))
        except (TypeError, ValueError):
            return 20

    def parse_distribution_column_px(self):
        app = self.app
        try:
            return float(str(app.graph_distribution_column_px_var.get()).strip())
        except (TypeError, ValueError):
            return 0.0

    def distribution_kind_defaults(self):
        app = self.app
        kind = str(app.graph_distribution_kind_var.get()).strip()
        unit_label = self.profile_axis_unit_label()
        if kind == "Positions":
            default_value_label = app.graph_y_axis_label.get().strip() or f"Vertical Position ({unit_label})"
        elif kind == "Z-Scores":
            # Z-scores are unitless; keep the axis label explicit and stable.
            default_value_label = "Z-Score"
            unit_label = "z"
        else:
            default_value_label = app.graph_y_axis_label.get().strip() or f"Residual ({unit_label})"
        return kind, unit_label, default_value_label

    def transform_distribution_values(self, raw_px, kind):
        raw_px = np.asarray(raw_px, dtype=np.float64)
        raw_px = raw_px[np.isfinite(raw_px)]
        if raw_px.size < 3:
            return None

        raw_mean_px = float(np.mean(raw_px))
        raw_std_px = float(np.std(raw_px))

        if kind == "Positions":
            values = np.array([float(self.profile_y_value(value)) for value in raw_px], dtype=np.float64)
        elif kind == "Z-Scores":
            if raw_std_px <= 0:
                values = np.zeros(raw_px.shape[0], dtype=np.float64)
            else:
                values = (raw_px - raw_mean_px) / raw_std_px
        else:
            residual_px = raw_px - raw_mean_px
            values = np.array([float(self.profile_y_delta_value(value)) for value in residual_px], dtype=np.float64)

        values = values[np.isfinite(values)]
        if values.size < 3:
            return None
        return values

    def build_histogram_across_columns_data(self, samples):
        app = self.app
        kind, unit_label, default_value_label = self.distribution_kind_defaults()
        bins = self.parse_distribution_bins()

        column_count = samples.shape[1]
        x_column_px = np.array([self.profile_index_to_x_px(idx) for idx in range(column_count)], dtype=np.float64)
        x_column_units = np.array([self.x_position_to_graph_units(value) for value in x_column_px], dtype=np.float64)
        x_column_display = np.array([self.column_index_to_value(idx) for idx in range(column_count)], dtype=np.float64)
        x_suffix = self.column_input_suffix()

        transformed_columns = []
        valid_mask = np.zeros(column_count, dtype=bool)
        collected_values = []
        for idx in range(column_count):
            transformed = self.transform_distribution_values(samples[:, idx], kind)
            transformed_columns.append(transformed)
            if transformed is not None:
                valid_mask[idx] = True
                collected_values.append(transformed)

        if not collected_values:
            return None

        all_values = np.concatenate(collected_values)
        min_value = float(np.min(all_values))
        max_value = float(np.max(all_values))
        if max_value <= min_value:
            min_value -= 0.5
            max_value += 0.5

        histogram_matrix = np.zeros((column_count, bins), dtype=np.int32)
        histogram_edges = None
        per_column_mean = np.full(column_count, np.nan, dtype=np.float64)
        per_column_std = np.full(column_count, np.nan, dtype=np.float64)

        for idx, values in enumerate(transformed_columns):
            if values is None:
                continue
            counts, edges = np.histogram(values, bins=bins, range=(min_value, max_value))
            histogram_matrix[idx, :] = counts
            histogram_edges = edges
            per_column_mean[idx] = float(np.mean(values))
            per_column_std[idx] = float(np.std(values))

        if histogram_edges is None:
            return None

        valid_column_indices = np.where(valid_mask)[0]
        median_std = float(np.nanmedian(per_column_std[valid_mask])) if np.any(valid_mask) else 0.0

        return {
            "mode": "Histogram",
            "kind": kind,
            "all_columns": True,
            "unit_label": unit_label,
            "column_input_mode": self.column_input_mode(),
            "column_input_suffix": x_suffix,
            "default_x_label": app.graph_x_axis_label.get().strip() or f"Horizontal Position ({x_suffix})",
            "default_y_label": default_value_label,
            "histogram_matrix": histogram_matrix,
            "histogram_edges": histogram_edges,
            "x_column_px": x_column_px,
            "x_column_units": x_column_units,
            "x_column_display": x_column_display,
            "valid_column_mask": valid_mask,
            "valid_column_indices": valid_column_indices,
            "sample_count": int(all_values.size),
            "column_count": int(column_count),
            "valid_column_count": int(valid_column_indices.size),
            "median_std": median_std,
            "per_column_mean": per_column_mean,
            "per_column_std": per_column_std,
        }

    def build_histogram_combined_columns_data(self, samples):
        app = self.app
        kind, unit_label, default_value_label = self.distribution_kind_defaults()
        bins = self.parse_distribution_bins()

        column_count = samples.shape[1]
        pooled_values = []
        valid_columns = 0
        for idx in range(column_count):
            transformed = self.transform_distribution_values(samples[:, idx], kind)
            if transformed is None:
                continue
            valid_columns += 1
            pooled_values.append(transformed)

        if not pooled_values:
            return None

        values = np.concatenate(pooled_values)
        values = values[np.isfinite(values)]
        if values.size < 3:
            return None

        value_mean = float(np.mean(values))
        value_std = float(np.std(values))
        centered = values - value_mean
        variance = float(np.mean(centered ** 2))
        if variance > 0:
            skewness = float(np.mean(centered ** 3) / (variance ** 1.5))
            excess_kurtosis = float(np.mean(centered ** 4) / (variance ** 2) - 3.0)
        else:
            skewness = 0.0
            excess_kurtosis = 0.0

        percentiles = np.percentile(values, [10, 50, 90])
        sorted_values = np.sort(values)
        probabilities = (np.arange(1, sorted_values.size + 1, dtype=np.float64) - 0.5) / sorted_values.size
        if value_std > 0:
            normal_dist = NormalDist(mu=value_mean, sigma=value_std)
            theoretical_quantiles = np.array([normal_dist.inv_cdf(float(prob)) for prob in probabilities], dtype=np.float64)
        else:
            theoretical_quantiles = np.full(sorted_values.shape, value_mean, dtype=np.float64)

        min_value = float(np.min(values))
        max_value = float(np.max(values))
        if max_value <= min_value:
            min_value -= 0.5
            max_value += 0.5
        histogram_counts, histogram_edges = np.histogram(values, bins=bins, range=(min_value, max_value))

        return {
            "mode": "Histogram",
            "kind": kind,
            "combined_columns": True,
            "unit_label": unit_label,
            "default_x_label": default_value_label,
            "default_y_label": "Sample Count",
            "values": values,
            "mean": value_mean,
            "std": value_std,
            "skewness": skewness,
            "excess_kurtosis": excess_kurtosis,
            "percentiles": percentiles,
            "sorted_values": sorted_values,
            "theoretical_quantiles": theoretical_quantiles,
            "histogram_counts": histogram_counts,
            "histogram_edges": histogram_edges,
            "sample_count": int(values.size),
            "column_count": int(column_count),
            "valid_column_count": int(valid_columns),
        }

    def selected_distribution_index(self, samples):
        if samples.ndim != 2 or samples.shape[1] <= 0:
            return None

        sample_width = samples.shape[1]
        requested_index = self.column_value_to_index(self.parse_distribution_column_px(), sample_width)

        def sync_entry_to_index(index_to_show):
            try:
                active_widget = self.app.root.focus_get()
                if active_widget is not self.app.graph_distribution_column_entry:
                    self.app.graph_distribution_column_px_var.set(self.format_graph_value(self.column_index_to_value(index_to_show)))
            except Exception:
                pass

        valid_counts = np.sum(np.isfinite(samples), axis=0)
        if valid_counts[requested_index] >= 3:
            sync_entry_to_index(requested_index)
            return requested_index

        candidate_indices = np.where(valid_counts >= 3)[0]
        if candidate_indices.size == 0:
            return None

        nearest = int(candidate_indices[np.argmin(np.abs(candidate_indices - requested_index))])
        sync_entry_to_index(nearest)
        return nearest

    def build_distribution_data(self):
        app = self.app
        if app.final_centerline_samples is None:
            imported_distribution = getattr(app, "imported_distribution_data", None)
            if not imported_distribution:
                return None

            kind = str(app.graph_distribution_kind_var.get()).strip()
            if kind == "Positions":
                values = np.asarray(imported_distribution.get("position_values", []), dtype=np.float64)
                unit_label = str(imported_distribution.get("unit_label", "")).strip() or "value"
            elif kind == "Z-Scores":
                values = np.asarray(imported_distribution.get("z_score_values", []), dtype=np.float64)
                unit_label = "z"
            else:
                values = np.asarray(imported_distribution.get("residual_values", []), dtype=np.float64)
                unit_label = str(imported_distribution.get("unit_label", "")).strip() or "value"

            values = values[np.isfinite(values)]
            if values.size < 3:
                return None

            value_mean = float(np.mean(values))
            value_std = float(np.std(values))
            centered = values - value_mean
            variance = float(np.mean(centered ** 2))
            if variance > 0:
                skewness = float(np.mean(centered ** 3) / (variance ** 1.5))
                excess_kurtosis = float(np.mean(centered ** 4) / (variance ** 2) - 3.0)
            else:
                skewness = 0.0
                excess_kurtosis = 0.0

            percentiles = np.percentile(values, [10, 50, 90])
            sorted_values = np.sort(values)
            probabilities = (np.arange(1, sorted_values.size + 1, dtype=np.float64) - 0.5) / sorted_values.size
            if value_std > 0:
                normal_dist = NormalDist(mu=value_mean, sigma=value_std)
                theoretical_quantiles = np.array([normal_dist.inv_cdf(float(prob)) for prob in probabilities], dtype=np.float64)
            else:
                theoretical_quantiles = np.full(sorted_values.shape, value_mean, dtype=np.float64)

            bins = self.parse_distribution_bins()
            min_value = float(np.min(values))
            max_value = float(np.max(values))
            if max_value <= min_value:
                min_value -= 0.5
                max_value += 0.5
            histogram_counts, histogram_edges = np.histogram(values, bins=bins, range=(min_value, max_value))

            return {
                "mode": str(app.graph_view_mode_var.get()).strip(),
                "kind": kind,
                "combined_columns": True,
                "imported": True,
                "unit_label": unit_label,
                "default_x_label": self._distribution_default_x_label(kind, unit_label),
                "default_y_label": "Sample Count",
                "values": values,
                "mean": value_mean,
                "std": value_std,
                "skewness": skewness,
                "excess_kurtosis": excess_kurtosis,
                "percentiles": percentiles,
                "sorted_values": sorted_values,
                "theoretical_quantiles": theoretical_quantiles,
                "histogram_counts": histogram_counts,
                "histogram_edges": histogram_edges,
                "sample_count": int(values.size),
                "column_count": int(imported_distribution.get("column_count", 0)),
                "valid_column_count": int(imported_distribution.get("valid_column_count", 0)),
            }

        samples = np.asarray(app.final_centerline_samples, dtype=np.float64)
        if samples.ndim != 2 or samples.shape[0] == 0 or samples.shape[1] == 0:
            return None

        mode = str(app.graph_view_mode_var.get()).strip()
        if mode == "Histogram":
            histogram_scope_var = getattr(app, "graph_histogram_scope_var", None)
            histogram_scope = str(histogram_scope_var.get()).strip() if histogram_scope_var is not None else "All Columns"
            if histogram_scope == "All Columns":
                return self.build_histogram_across_columns_data(samples)
            if histogram_scope == "All Columns (Combined)":
                return self.build_histogram_combined_columns_data(samples)
        elif mode == "Q-Q Plot":
            combined_data = self.build_histogram_combined_columns_data(samples)
            if combined_data is None:
                return None
            combined_data["mode"] = "Q-Q Plot"
            return combined_data

        selected_index = self.selected_distribution_index(samples)
        if selected_index is None:
            return None
        raw_px = samples[:, selected_index]

        values = self.transform_distribution_values(raw_px, str(app.graph_distribution_kind_var.get()).strip())
        if values is None:
            return None

        selected_x_px = float(self.profile_index_to_x_px(selected_index))
        selected_x_units = float(self.x_position_to_graph_units(selected_x_px))
        selected_x_display = float(self.column_index_to_value(selected_index))
        kind, unit_label, default_value_label = self.distribution_kind_defaults()

        value_mean = float(np.mean(values))
        value_std = float(np.std(values))
        centered = values - value_mean
        variance = float(np.mean(centered ** 2))
        if variance > 0:
            skewness = float(np.mean(centered ** 3) / (variance ** 1.5))
            excess_kurtosis = float(np.mean(centered ** 4) / (variance ** 2) - 3.0)
        else:
            skewness = 0.0
            excess_kurtosis = 0.0

        percentiles = np.percentile(values, [10, 50, 90])
        sorted_values = np.sort(values)
        probabilities = (np.arange(1, sorted_values.size + 1, dtype=np.float64) - 0.5) / sorted_values.size
        if value_std > 0:
            normal_dist = NormalDist(mu=value_mean, sigma=value_std)
            theoretical_quantiles = np.array([normal_dist.inv_cdf(float(prob)) for prob in probabilities], dtype=np.float64)
        else:
            theoretical_quantiles = np.full(sorted_values.shape, value_mean, dtype=np.float64)

        bins = self.parse_distribution_bins()
        min_value = float(np.min(values))
        max_value = float(np.max(values))
        if max_value <= min_value:
            min_value -= 0.5
            max_value += 0.5
        histogram_counts, histogram_edges = np.histogram(values, bins=bins, range=(min_value, max_value))

        return {
            "mode": str(app.graph_view_mode_var.get()).strip(),
            "kind": kind,
            "values": values,
            "selected_index": selected_index,
            "selected_x_px": selected_x_px,
            "selected_x_units": selected_x_units,
            "selected_x_display": selected_x_display,
            "column_input_mode": self.column_input_mode(),
            "column_input_suffix": self.column_input_suffix(),
            "unit_label": unit_label,
            "default_x_label": default_value_label,
            "default_y_label": "Sample Count",
            "mean": value_mean,
            "std": value_std,
            "skewness": skewness,
            "excess_kurtosis": excess_kurtosis,
            "percentiles": percentiles,
            "sorted_values": sorted_values,
            "theoretical_quantiles": theoretical_quantiles,
            "histogram_counts": histogram_counts,
            "histogram_edges": histogram_edges,
        }

    def build_distribution_summary(self, distribution_data):
        if distribution_data.get("all_columns"):
            return (
                f"All columns histogram | valid columns={distribution_data['valid_column_count']}/{distribution_data['column_count']} "
                f"| total samples={distribution_data['sample_count']}\n"
                f"median column std={distribution_data['median_std']:.4g} ({distribution_data['kind']})"
            )

        if distribution_data.get("combined_columns"):
            mean_value = distribution_data["mean"]
            std_value = distribution_data["std"]
            skewness = distribution_data["skewness"]
            excess_kurtosis = distribution_data["excess_kurtosis"]
            p10, p50, p90 = distribution_data["percentiles"]
            return (
                f"Combined all-columns histogram | valid columns={distribution_data['valid_column_count']}/{distribution_data['column_count']} "
                f"| n={distribution_data['sample_count']}\n"
                f"mean={mean_value:.4g} | std={std_value:.4g} | skew={skewness:.4g} | excess kurtosis={excess_kurtosis:.4g}\n"
                f"p10={p10:.4g} | median={p50:.4g} | p90={p90:.4g}"
            )

        selected_x_px = distribution_data["selected_x_px"]
        selected_x_display = distribution_data.get("selected_x_display", selected_x_px)
        selected_suffix = distribution_data.get("column_input_suffix", "px")
        selected_index = distribution_data["selected_index"]
        mean_value = distribution_data["mean"]
        std_value = distribution_data["std"]
        skewness = distribution_data["skewness"]
        excess_kurtosis = distribution_data["excess_kurtosis"]
        p10, p50, p90 = distribution_data["percentiles"]
        return (
            f"Column {selected_x_display:.3g} {selected_suffix} (sample {selected_index}, pixel x={selected_x_px:.3g}px) | "
            f"n={distribution_data['values'].size} | mean={mean_value:.4g} | std={std_value:.4g}\n"
            f"skew={skewness:.4g} | excess kurtosis={excess_kurtosis:.4g}\n"
            f"p10={p10:.4g} | median={p50:.4g} | p90={p90:.4g}"
        )

    def redraw_graph(self):
        self.refresh_distribution_column_controls()
        mode = str(self.app.graph_view_mode_var.get()).strip()
        if mode == "Profile":
            self.redraw_profile_graph()
        else:
            self.redraw_distribution_graph(mode)

    def redraw_profile_graph(self):
        app = self.app
        app.graph_canvas.delete("all")
        app.graph_canvas.configure(cursor="")
        w = app.graph_canvas.winfo_width()
        h = app.graph_canvas.winfo_height()
        if w < 50 or h < 50:
            return

        plot_data = self.build_plot_data()
        if plot_data is None:
            app.set_graph_fit_equation_text("Best fit: n/a")
            app.graph_canvas.create_text(
                w // 2,
                h // 2,
                text="Run analysis or import a CSV to generate a graph.",
                fill="gray35",
                font=("TkDefaultFont", 12),
            )
            return

        source = str(plot_data.get("source", "native")).strip()
        valid = plot_data["valid"]
        stdevs = plot_data["stdevs"]

        if not np.any(valid):
            app.set_graph_fit_equation_text("Best fit: n/a")
            app.graph_canvas.create_text(
                w // 2,
                h // 2,
                text="No valid centerline data to plot.",
                fill="gray35",
                font=("TkDefaultFont", 12),
            )
            return

        valid_idx = np.where(valid)[0]
        x_all_values = plot_data["x_all_values"]
        mean_values = plot_data["mean_values"]
        upper_values = plot_data["upper_values"]
        lower_values = plot_data["lower_values"]
        y_range_values = np.concatenate((upper_values[valid], lower_values[valid]))
        limits = self.resolve_axis_limits(x_all_values[valid], y_range_values, y_pad=1.0)
        if limits is None:
            app.set_graph_fit_equation_text("Best fit: n/a")
            return
        x_min, x_max, y_min, y_max = limits

        left = 110
        right = 20
        top = 28
        bottom = 62
        plot_w = max(1, w - left - right)
        plot_h = max(1, h - top - bottom)

        def x_to_px(x_value):
            if x_max <= x_min:
                return left + (plot_w / 2.0)
            return left + ((x_value - x_min) / (x_max - x_min)) * plot_w

        def y_to_px(y_value):
            return top + ((y_max - y_value) / (y_max - y_min)) * plot_h

        app.graph_canvas.create_rectangle(left, top, left + plot_w, top + plot_h, fill="white", outline="")
        app.graph_canvas.create_line(left, top + plot_h, left + plot_w, top + plot_h, fill="black")
        app.graph_canvas.create_line(left, top, left, top + plot_h, fill="black")

        y_ticks = 5
        for i in range(y_ticks + 1):
            frac = i / y_ticks
            y_val = y_min + (y_max - y_min) * frac
            y_px = y_to_px(y_val)
            app.graph_canvas.create_line(left - 5, y_px, left, y_px, fill="black")
            y_tags = ()
            if i == 0:
                y_tags = ("graph_edit_y_min_tick",)
            elif i == y_ticks:
                y_tags = ("graph_edit_y_max_tick",)
            app.graph_canvas.create_text(
                left - 8,
                y_px,
                text=self.format_graph_value(y_val),
                anchor="e",
                fill="black",
                font=("TkDefaultFont", 9),
                tags=y_tags,
            )

        x_ticks = 6
        for i in range(x_ticks):
            frac = i / max(1, (x_ticks - 1))
            x_val = x_min + frac * (x_max - x_min)
            x_px = x_to_px(x_val)
            app.graph_canvas.create_line(x_px, top + plot_h, x_px, top + plot_h + 5, fill="black")
            x_tags = ()
            if i == 0:
                x_tags = ("graph_edit_x_min_tick",)
            elif i == (x_ticks - 1):
                x_tags = ("graph_edit_x_max_tick",)
            app.graph_canvas.create_text(
                x_px,
                top + plot_h + 16,
                text=self.format_graph_value(x_val),
                fill="black",
                font=("TkDefaultFont", 9),
                tags=x_tags,
            )

        band_points = []
        for idx in valid_idx:
            band_points.extend((x_to_px(x_all_values[idx]), y_to_px(upper_values[idx])))
        for idx in valid_idx[::-1]:
            band_points.extend((x_to_px(x_all_values[idx]), y_to_px(lower_values[idx])))
        if len(band_points) >= 6:
            app.graph_canvas.create_polygon(band_points, fill="#cfe8ff", outline="")

        mean_points = []
        for idx in valid_idx:
            mean_points.extend((x_to_px(x_all_values[idx]), y_to_px(mean_values[idx])))
        if len(mean_points) >= 4:
            app.graph_canvas.create_line(mean_points, fill="#005fbd", width=2)

        if app.show_best_fit_var.get():
            fit = self.compute_plot_fit(plot_data)
            if fit is not None:
                fit_points = []
                for idx in valid_idx:
                    if source == "imported_csv":
                        fit_y_value = float(fit["y_fit_values"][idx])
                    else:
                        fit_y_value = float(self.profile_y_value(fit["y_fit"][idx]))
                    fit_points.extend((x_to_px(x_all_values[idx]), y_to_px(fit_y_value)))
                if len(fit_points) >= 4:
                    app.graph_canvas.create_line(fit_points, fill="#d62728", width=2, dash=(6, 4))
                app.set_graph_fit_equation_text(fit["equation"])
            else:
                app.set_graph_fit_equation_text("Best fit: n/a")
        else:
            app.set_graph_fit_equation_text("Best fit: hidden")

        app.graph_canvas.create_text(
            w // 2,
            10,
            text=self.resolve_graph_title(plot_data.get("title", f"Final Mean Centerline with +/- {stdevs:g}sigma Band ({self.profile_axis_unit_label()})")),
            fill="black",
            anchor="n",
            font=("TkDefaultFont", 11, "bold"),
            tags=("graph_edit_title",),
        )
        self._bind_graph_edit_action("graph_edit_title", self.edit_graph_title)
        self._draw_editable_axis_controls(left, top, plot_w, plot_h)

    def redraw_distribution_graph(self, mode):
        app = self.app
        app.graph_canvas.delete("all")
        app.graph_canvas.configure(cursor="")
        w = app.graph_canvas.winfo_width()
        h = app.graph_canvas.winfo_height()
        if w < 50 or h < 50:
            return

        distribution_data = self.build_distribution_data()
        if distribution_data is None:
            app.set_graph_fit_equation_text("Distribution data: n/a")
            app.graph_canvas.create_text(
                w // 2,
                h // 2,
                text="Run video analysis or import a compatible graph CSV to view histogram and Q-Q data.",
                fill="gray35",
                font=("TkDefaultFont", 12),
                width=max(200, w - 80),
            )
            return

        app.set_graph_fit_equation_text(self.build_distribution_summary(distribution_data))

        left = 110
        right = 20
        top = 28
        bottom = 62
        plot_w = max(1, w - left - right)
        plot_h = max(1, h - top - bottom)

        if mode == "Histogram":
            if distribution_data.get("all_columns"):
                edges = distribution_data["histogram_edges"]
                counts_matrix = distribution_data["histogram_matrix"]
                x_column_display = distribution_data["x_column_display"]
                max_count = int(np.max(counts_matrix)) if counts_matrix.size > 0 else 0
                if max_count <= 0:
                    return

                if x_column_display.size == 1:
                    column_span = 1.0
                    x_edges = np.array([x_column_display[0] - 0.5 * column_span, x_column_display[0] + 0.5 * column_span], dtype=np.float64)
                else:
                    midpoints = 0.5 * (x_column_display[:-1] + x_column_display[1:])
                    first_edge = x_column_display[0] - (midpoints[0] - x_column_display[0])
                    last_edge = x_column_display[-1] + (x_column_display[-1] - midpoints[-1])
                    x_edges = np.concatenate(([first_edge], midpoints, [last_edge]))

                x_min = float(x_edges[0])
                x_max = float(x_edges[-1])
                y_min = float(edges[0])
                y_max = float(edges[-1])

                def x_to_px(x_value):
                    return left + ((x_value - x_min) / (x_max - x_min)) * plot_w if x_max > x_min else left + (plot_w / 2.0)

                def y_to_px(y_value):
                    return top + ((y_max - y_value) / max(y_max - y_min, 1.0)) * plot_h

                def heat_color(norm_value):
                    norm = max(0.0, min(1.0, float(norm_value)))
                    red = int(round(248 - 184 * norm))
                    green = int(round(251 - 196 * norm))
                    blue = int(round(255 - 21 * norm))
                    return f"#{red:02x}{green:02x}{blue:02x}"

                app.graph_canvas.create_rectangle(left, top, left + plot_w, top + plot_h, fill="white", outline="")
                app.graph_canvas.create_line(left, top + plot_h, left + plot_w, top + plot_h, fill="black")
                app.graph_canvas.create_line(left, top, left, top + plot_h, fill="black")

                for col_idx in range(counts_matrix.shape[0]):
                    x0 = x_to_px(float(x_edges[col_idx]))
                    x1 = x_to_px(float(x_edges[col_idx + 1]))
                    for bin_idx in range(counts_matrix.shape[1]):
                        count = int(counts_matrix[col_idx, bin_idx])
                        if count <= 0:
                            continue
                        y0 = y_to_px(float(edges[bin_idx + 1]))
                        y1 = y_to_px(float(edges[bin_idx]))
                        app.graph_canvas.create_rectangle(
                            x0,
                            y0,
                            x1,
                            y1,
                            fill=heat_color(count / max_count),
                            outline="",
                        )

                y_ticks = 5
                for i in range(y_ticks + 1):
                    frac = i / y_ticks
                    y_val = y_min + (y_max - y_min) * frac
                    y_px = y_to_px(y_val)
                    app.graph_canvas.create_line(left - 5, y_px, left, y_px, fill="black")
                    y_tags = ()
                    if i == 0:
                        y_tags = ("graph_edit_y_min_tick",)
                    elif i == y_ticks:
                        y_tags = ("graph_edit_y_max_tick",)
                    app.graph_canvas.create_text(left - 8, y_px, text=self.format_graph_value(y_val), anchor="e", fill="black", font=("TkDefaultFont", 9), tags=y_tags)

                x_ticks = 6
                for i in range(x_ticks):
                    frac = i / max(1, (x_ticks - 1))
                    x_val = x_min + frac * (x_max - x_min)
                    x_px = x_to_px(x_val)
                    app.graph_canvas.create_line(x_px, top + plot_h, x_px, top + plot_h + 5, fill="black")
                    x_tags = ()
                    if i == 0:
                        x_tags = ("graph_edit_x_min_tick",)
                    elif i == (x_ticks - 1):
                        x_tags = ("graph_edit_x_max_tick",)
                    app.graph_canvas.create_text(x_px, top + plot_h + 16, text=self.format_graph_value(x_val), fill="black", font=("TkDefaultFont", 9), tags=x_tags)

                title = (
                    f"Histogram Heatmap of {distribution_data['kind']} Across Columns "
                    f"({distribution_data['sample_count']} samples)"
                )
                x_axis_text = distribution_data["default_x_label"]
                y_axis_text = distribution_data["default_y_label"]
                app.graph_canvas.create_text(left + 8, top + 12, text="Darker color = higher count", anchor="w", fill="gray30", font=("TkDefaultFont", 9))
            else:
                edges = distribution_data["histogram_edges"]
                counts = distribution_data["histogram_counts"]
                limits = self.resolve_axis_limits(edges, np.append(counts, 0), y_pad=1.0)
                if limits is None:
                    return
                x_min, x_max, _y_min, y_max = limits
                y_min = 0.0

                def x_to_px(x_value):
                    return left + ((x_value - x_min) / (x_max - x_min)) * plot_w if x_max > x_min else left + (plot_w / 2.0)

                def y_to_px(y_value):
                    return top + ((y_max - y_value) / max(y_max - y_min, 1.0)) * plot_h

                app.graph_canvas.create_rectangle(left, top, left + plot_w, top + plot_h, fill="white", outline="")
                app.graph_canvas.create_line(left, top + plot_h, left + plot_w, top + plot_h, fill="black")
                app.graph_canvas.create_line(left, top, left, top + plot_h, fill="black")

                for idx, count in enumerate(counts):
                    x0 = x_to_px(edges[idx])
                    x1 = x_to_px(edges[idx + 1])
                    y1 = y_to_px(float(count))
                    app.graph_canvas.create_rectangle(x0, y1, x1, top + plot_h, fill="#cfe8ff", outline="#4a7ebb")

                y_ticks = 5
                for i in range(y_ticks + 1):
                    frac = i / y_ticks
                    y_val = y_min + (y_max - y_min) * frac
                    y_px = y_to_px(y_val)
                    app.graph_canvas.create_line(left - 5, y_px, left, y_px, fill="black")
                    y_tags = ()
                    if i == 0:
                        y_tags = ("graph_edit_y_min_tick",)
                    elif i == y_ticks:
                        y_tags = ("graph_edit_y_max_tick",)
                    app.graph_canvas.create_text(left - 8, y_px, text=self.format_graph_value(y_val), anchor="e", fill="black", font=("TkDefaultFont", 9), tags=y_tags)

                x_ticks = 6
                for i in range(x_ticks):
                    frac = i / max(1, (x_ticks - 1))
                    x_val = x_min + frac * (x_max - x_min)
                    x_px = x_to_px(x_val)
                    app.graph_canvas.create_line(x_px, top + plot_h, x_px, top + plot_h + 5, fill="black")
                    x_tags = ()
                    if i == 0:
                        x_tags = ("graph_edit_x_min_tick",)
                    elif i == (x_ticks - 1):
                        x_tags = ("graph_edit_x_max_tick",)
                    app.graph_canvas.create_text(x_px, top + plot_h + 16, text=self.format_graph_value(x_val), fill="black", font=("TkDefaultFont", 9), tags=x_tags)

                std_value = distribution_data["std"]
                if std_value > 0:
                    x_curve = np.linspace(x_min, x_max, 200)
                    bin_width = edges[1] - edges[0] if len(edges) > 1 else 1.0
                    normal_curve = np.array([
                        distribution_data["values"].size * bin_width * NormalDist(mu=distribution_data["mean"], sigma=std_value).pdf(float(x_val))
                        for x_val in x_curve
                    ], dtype=np.float64)
                    curve_points = []
                    for x_val, y_val in zip(x_curve, normal_curve):
                        curve_points.extend((x_to_px(float(x_val)), y_to_px(float(y_val))))
                    if len(curve_points) >= 4:
                        app.graph_canvas.create_line(curve_points, fill="#d62728", width=2, smooth=True)

                if distribution_data.get("combined_columns"):
                    title = (
                        f"Histogram of {distribution_data['kind']} Across All Columns "
                        f"({distribution_data['values'].size} samples)"
                    )
                else:
                    title = (
                        f"Histogram of {distribution_data['kind']} at x={distribution_data['selected_x_display']:.3g} "
                        f"{distribution_data.get('column_input_suffix', 'px')} "
                        f"({distribution_data['values'].size} samples)"
                    )
                x_axis_text = distribution_data["default_x_label"]
                y_axis_text = distribution_data["default_y_label"]
        else:
            theoretical = distribution_data["theoretical_quantiles"]
            observed = distribution_data["sorted_values"]
            combined = np.concatenate((theoretical, observed))
            limits = self.resolve_axis_limits(combined, combined, y_pad=1.0)
            if limits is None:
                return
            x_min, x_max, y_min, y_max = limits

            def x_to_px(x_value):
                return left + ((x_value - x_min) / (x_max - x_min)) * plot_w if x_max > x_min else left + (plot_w / 2.0)

            def y_to_px(y_value):
                return top + ((y_max - y_value) / max(y_max - y_min, 1.0)) * plot_h

            app.graph_canvas.create_rectangle(left, top, left + plot_w, top + plot_h, fill="white", outline="")
            app.graph_canvas.create_line(left, top + plot_h, left + plot_w, top + plot_h, fill="black")
            app.graph_canvas.create_line(left, top, left, top + plot_h, fill="black")

            diag_start_x = max(x_min, y_min)
            diag_end_x = min(x_max, y_max)
            if diag_end_x > diag_start_x:
                app.graph_canvas.create_line(
                    x_to_px(diag_start_x),
                    y_to_px(diag_start_x),
                    x_to_px(diag_end_x),
                    y_to_px(diag_end_x),
                    fill="#999999",
                    dash=(4, 3),
                )

            for x_val, y_val in zip(theoretical, observed):
                x_px = x_to_px(float(x_val))
                y_px = y_to_px(float(y_val))
                app.graph_canvas.create_oval(x_px - 2, y_px - 2, x_px + 2, y_px + 2, fill="#005fbd", outline="")

            y_ticks = 5
            for i in range(y_ticks + 1):
                frac = i / y_ticks
                y_val = y_min + (y_max - y_min) * frac
                y_px = y_to_px(y_val)
                app.graph_canvas.create_line(left - 5, y_px, left, y_px, fill="black")
                y_tags = ()
                if i == 0:
                    y_tags = ("graph_edit_y_min_tick",)
                elif i == y_ticks:
                    y_tags = ("graph_edit_y_max_tick",)
                app.graph_canvas.create_text(left - 8, y_px, text=self.format_graph_value(y_val), anchor="e", fill="black", font=("TkDefaultFont", 9), tags=y_tags)

            x_ticks = 6
            for i in range(x_ticks):
                frac = i / max(1, (x_ticks - 1))
                x_val = x_min + frac * (x_max - x_min)
                x_px = x_to_px(x_val)
                app.graph_canvas.create_line(x_px, top + plot_h, x_px, top + plot_h + 5, fill="black")
                x_tags = ()
                if i == 0:
                    x_tags = ("graph_edit_x_min_tick",)
                elif i == (x_ticks - 1):
                    x_tags = ("graph_edit_x_max_tick",)
                app.graph_canvas.create_text(x_px, top + plot_h + 16, text=self.format_graph_value(x_val), fill="black", font=("TkDefaultFont", 9), tags=x_tags)

            if distribution_data.get("combined_columns"):
                title = f"Normal Q-Q Plot of {distribution_data['kind']} Across All Columns"
            else:
                title = (
                    f"Normal Q-Q Plot of {distribution_data['kind']} at x={distribution_data['selected_x_display']:.3g} "
                    f"{distribution_data.get('column_input_suffix', 'px')}"
                )
            x_axis_text = app.graph_x_axis_label.get().strip() or "Theoretical Normal Quantile"
            y_axis_text = app.graph_y_axis_label.get().strip() or "Observed Quantile"

        app.graph_canvas.create_text(
            w // 2,
            10,
            text=self.resolve_graph_title(title),
            fill="black",
            anchor="n",
            font=("TkDefaultFont", 11, "bold"),
            tags=("graph_edit_title",),
        )
        self._bind_graph_edit_action("graph_edit_title", self.edit_graph_title)
        # Keep vars synced with the currently active graph labels so canvas editing always targets visible labels.
        if not app.graph_x_axis_label.get().strip():
            app.graph_x_axis_label.set(x_axis_text)
        if not app.graph_y_axis_label.get().strip():
            app.graph_y_axis_label.set(y_axis_text)
        self._draw_editable_axis_controls(left, top, plot_w, plot_h)

    def build_plot_data(self):
        app = self.app
        imported = getattr(app, "imported_profile_data", None)
        if imported:
            x_all_values = np.asarray(imported.get("x_all_values", []), dtype=np.float64)
            mean_values = np.asarray(imported.get("mean_values", []), dtype=np.float64)
            upper_values = np.asarray(imported.get("upper_values", []), dtype=np.float64)
            lower_values = np.asarray(imported.get("lower_values", []), dtype=np.float64)
            if x_all_values.size == 0 or mean_values.size == 0:
                return None
            valid = np.asarray(imported.get("valid", np.isfinite(x_all_values) & np.isfinite(mean_values)), dtype=bool)
            if not np.any(valid):
                return None
            title = str(imported.get("title", "")).strip()
            unit_label = str(imported.get("unit_label", "")).strip()
            return {
                "source": "imported_csv",
                "valid": valid,
                "stdevs": float(imported.get("stdevs", self.parse_graph_stdevs())),
                "x_all_values": x_all_values,
                "mean_values": mean_values,
                "upper_values": upper_values,
                "lower_values": lower_values,
                "title": title or f"Imported CSV Profile ({unit_label or 'values'})",
                "unit_label": unit_label or "values",
                "metadata": dict(imported.get("metadata", {})),
            }

        if app.final_mean_profile is None or app.final_std_profile is None:
            return None
        mean = np.asarray(app.final_mean_profile, dtype=np.float64)
        std = np.asarray(app.final_std_profile, dtype=np.float64)
        if mean.size == 0:
            return None
        stdevs = self.parse_graph_stdevs()
        safe_std = np.where(np.isfinite(std), std, 0.0)
        upper = mean + stdevs * safe_std
        lower = mean - stdevs * safe_std
        valid = np.isfinite(mean)
        if not np.any(valid):
            return None
        x_all_values = np.array(
            [self.profile_x_value(self.profile_index_to_x_px(i)) for i in range(mean.size)],
            dtype=np.float64,
        )
        mean_values = np.full(mean.shape, np.nan, dtype=np.float64)
        upper_values = np.full(mean.shape, np.nan, dtype=np.float64)
        lower_values = np.full(mean.shape, np.nan, dtype=np.float64)
        valid_idx = np.where(valid)[0]
        for idx in valid_idx:
            mean_values[idx] = float(self.profile_y_value(mean[idx]))
            upper_values[idx] = float(self.profile_y_value(upper[idx]))
            lower_values[idx] = float(self.profile_y_value(lower[idx]))
        return {
            "source": "native",
            "mean": mean,
            "std": std,
            "upper": upper,
            "lower": lower,
            "valid": valid,
            "stdevs": stdevs,
            "x_all_values": x_all_values,
            "mean_values": mean_values,
            "upper_values": upper_values,
            "lower_values": lower_values,
        }

    def build_graph_export_rows(self):
        app = self.app
        plot_data = self.build_plot_data()
        if plot_data is None:
            return None

        if str(plot_data.get("source", "native")).strip() == "imported_csv":
            fit = self.compute_plot_fit(plot_data)
            rows = []
            valid = plot_data["valid"]
            x_all_values = plot_data["x_all_values"]
            mean_values = plot_data["mean_values"]
            lower_values = plot_data["lower_values"]
            upper_values = plot_data["upper_values"]
            fit_values = fit["y_fit_values"] if fit is not None else None
            for idx in range(x_all_values.size):
                if not bool(valid[idx]):
                    continue
                rows.append([
                    float(x_all_values[idx]),
                    float(mean_values[idx]),
                    float(lower_values[idx]),
                    float(upper_values[idx]),
                    float(fit_values[idx]) if fit_values is not None and np.isfinite(fit_values[idx]) else "",
                ])

            metadata = [["export_type", "graph_profile_imported_csv"]]
            metadata.extend([[str(key), str(value)] for key, value in plot_data.get("metadata", {}).items()])
            metadata.extend([
                ["source_csv", str(getattr(app, "imported_profile_data", {}).get("source_path", ""))],
                ["fit_degree", str(fit["degree"]) if fit is not None else ""],
                ["fit_equation", fit["equation"] if fit is not None else "n/a"],
                ["generated_at", datetime.now().isoformat(timespec="seconds")],
            ])
            unit_label = str(plot_data.get("unit_label", "value")).strip() or "value"
            return {
                "metadata": metadata,
                "header": [
                    f"x_{unit_label}",
                    f"mean_y_{unit_label}",
                    f"lower_band_y_{unit_label}",
                    f"upper_band_y_{unit_label}",
                    f"fit_y_{unit_label}",
                ],
                "rows": rows,
            }

        mean = plot_data["mean"]
        upper = plot_data["upper"]
        lower = plot_data["lower"]
        valid = plot_data["valid"]
        stdevs = plot_data["stdevs"]
        std = np.asarray(plot_data["std"], dtype=np.float64)
        fit = self.compute_plot_fit(plot_data)

        unit_label = (app.graph_unit_label or "px").strip()
        unit_suffix = unit_label.lower()

        pixel_header = [
            "column_px",
            "mean_y_graph_px",
            "lower_band_y_graph_px",
            "upper_band_y_graph_px",
            "fit_y_graph_px",
        ]
        unit_header = [
            f"column_{unit_suffix}",
            f"mean_y_{unit_suffix}",
            f"lower_band_y_{unit_suffix}",
            f"upper_band_y_{unit_suffix}",
            f"fit_y_{unit_suffix}",
        ]

        rows = []
        for idx in range(mean.size):
            is_valid = bool(valid[idx])
            # Skip rows with no data
            if not is_valid:
                continue
            
            x_px = self.profile_index_to_x_px(idx)
            mean_raw_px = float(mean[idx])
            mean_px = self.y_position_to_graph_pixels(mean_raw_px)
            std_px = float(std[idx]) if np.isfinite(std[idx]) else 0.0
            lower_raw_px = float(lower[idx])
            upper_raw_px = float(upper[idx])
            lower_px = self.y_position_to_graph_pixels(lower_raw_px)
            upper_px = self.y_position_to_graph_pixels(upper_raw_px)
            mean_unit = float(self.y_position_to_graph_units(mean_raw_px))
            std_unit = float(self.y_delta_to_graph_units(std_px))
            lower_unit = float(self.y_position_to_graph_units(lower_raw_px))
            upper_unit = float(self.y_position_to_graph_units(upper_raw_px))
            if fit is not None:
                fit_raw_px = float(fit["y_fit"][idx])
                fit_px = self.y_position_to_graph_pixels(fit_raw_px)
                fit_unit = float(self.y_position_to_graph_units(fit_raw_px))
            else:
                fit_px = ""
                fit_unit = ""

            rows.append(
                {
                    "pixel": [
                        x_px,
                        mean_px,
                        lower_px,
                        upper_px,
                        fit_px,
                    ],
                    "unit": [
                        float(self.x_position_to_graph_units(x_px)),
                        mean_unit,
                        lower_unit,
                        upper_unit,
                        fit_unit,
                    ],
                }
            )

        metadata = [
            ["export_type", "graph_profile"],
            ["unit_label", unit_label],
            ["pixel_y_reference", "graph_oriented"],
            ["unit_scale", f"{app.graph_unit_scale:.12g}"],
            ["stdev_multiplier", f"{stdevs:g}"],
            ["fit_degree", str(fit["degree"]) if fit is not None else ""],
            ["fit_equation", fit["equation"] if fit is not None else "n/a"],
            ["nozzle_origin_y_px", f"{app.nozzle_origin_img[1]:.6g}" if app.nozzle_origin_img is not None else ""],
            ["source_video", app.video_path.get().strip()],
            ["generated_at", datetime.now().isoformat(timespec="seconds")],
        ]

        return {
            "metadata": metadata,
            "pixel_header": pixel_header,
            "unit_header": unit_header,
            "rows": rows,
        }

    def _build_combined_distribution_export(self, samples):
        """Always export combined-columns distribution with all three transform columns."""
        samples = np.asarray(samples, dtype=np.float64)
        if samples.ndim != 2 or samples.shape[0] == 0 or samples.shape[1] == 0:
            return None

        col_count = samples.shape[1]
        positions_list = []
        residuals_list = []
        zscores_list = []
        valid_columns = 0

        for col_idx in range(col_count):
            raw_px = samples[:, col_idx]
            raw_px = raw_px[np.isfinite(raw_px)]
            if raw_px.size < 3:
                continue

            raw_mean = float(np.mean(raw_px))
            raw_std = float(np.std(raw_px))

            pos_vals = np.array(
                [float(self.profile_y_value(v)) for v in raw_px], dtype=np.float64
            )
            residual_px = raw_px - raw_mean
            res_vals = np.array(
                [float(self.profile_y_delta_value(v)) for v in residual_px], dtype=np.float64
            )
            if raw_std > 0:
                z_vals = (raw_px - raw_mean) / raw_std
            else:
                z_vals = np.zeros_like(raw_px)

            # Use a common finite mask so all three arrays are aligned row-for-row
            finite_mask = np.isfinite(pos_vals) & np.isfinite(res_vals) & np.isfinite(z_vals)
            if np.sum(finite_mask) < 3:
                continue

            positions_list.append(pos_vals[finite_mask])
            residuals_list.append(res_vals[finite_mask])
            zscores_list.append(z_vals[finite_mask])
            valid_columns += 1

        if not positions_list:
            return None

        all_positions = np.concatenate(positions_list)
        all_residuals = np.concatenate(residuals_list)
        all_zscores = np.concatenate(zscores_list)
        n = all_zscores.size

        z_mean = float(np.mean(all_zscores))
        z_std = float(np.std(all_zscores))
        sorted_z = np.sort(all_zscores)
        probabilities = (np.arange(1, n + 1, dtype=np.float64) - 0.5) / n
        if z_std > 0:
            normal_dist = NormalDist(mu=z_mean, sigma=z_std)
            theoretical = np.array(
                [normal_dist.inv_cdf(float(p)) for p in probabilities], dtype=np.float64
            )
        else:
            theoretical = np.full(n, z_mean, dtype=np.float64)

        # Sort all columns by z-score order so the Q-Q theoretical column aligns
        sort_order = np.argsort(all_zscores)
        all_positions = all_positions[sort_order]
        all_residuals = all_residuals[sort_order]
        all_zscores = all_zscores[sort_order]

        rows = [
            [
                idx,
                float(all_positions[idx]),
                float(all_residuals[idx]),
                float(all_zscores[idx]),
                float(theoretical[idx]),
            ]
            for idx in range(n)
        ]

        z_centered = all_zscores - z_mean
        z_var = float(np.mean(z_centered ** 2))
        skewness = float(np.mean(z_centered ** 3) / (z_var ** 1.5)) if z_var > 0 else 0.0
        excess_kurtosis = float(np.mean(z_centered ** 4) / (z_var ** 2) - 3.0) if z_var > 0 else 0.0

        unit_label = (getattr(self.app, "graph_unit_label", None) or "px").strip()
        metadata = [
            ["export_type", "graph_distribution"],
            ["combined_columns", "true"],
            ["column_count", str(col_count)],
            ["valid_column_count", str(valid_columns)],
            ["sample_count", str(n)],
            ["z_score_mean", f"{z_mean:.12g}"],
            ["z_score_std", f"{z_std:.12g}"],
            ["z_score_skewness", f"{skewness:.12g}"],
            ["z_score_excess_kurtosis", f"{excess_kurtosis:.12g}"],
            ["position_unit", unit_label],
            ["residual_unit", unit_label],
            ["generated_at", datetime.now().isoformat(timespec="seconds")],
        ]

        return {
            "metadata": metadata,
            "header": [
                "sample_index",
                f"position_{unit_label}",
                f"residual_{unit_label}",
                "z_score",
                "theoretical_normal_quantile",
            ],
            "rows": rows,
        }

    def build_distribution_export_rows(self):
        distribution_data = self.build_distribution_data()
        if distribution_data is None:
            return None

        if distribution_data.get("all_columns"):
            rows = []
            histogram_matrix = distribution_data["histogram_matrix"]
            histogram_edges = distribution_data["histogram_edges"]
            x_column_px = distribution_data["x_column_px"]
            x_column_units = distribution_data["x_column_units"]
            for col_idx in range(histogram_matrix.shape[0]):
                for bin_idx in range(histogram_matrix.shape[1]):
                    rows.append([
                        col_idx,
                        float(x_column_px[col_idx]),
                        float(x_column_units[col_idx]),
                        float(histogram_edges[bin_idx]),
                        float(histogram_edges[bin_idx + 1]),
                        int(histogram_matrix[col_idx, bin_idx]),
                    ])

            metadata = [
                ["export_type", "graph_distribution_heatmap"],
                ["graph_view_mode", distribution_data["mode"]],
                ["distribution_kind", distribution_data["kind"]],
                ["column_input_mode", distribution_data.get("column_input_mode", "Pixel Values")],
                ["column_count", str(distribution_data["column_count"])],
                ["valid_column_count", str(distribution_data["valid_column_count"])],
                ["total_sample_count", str(distribution_data["sample_count"])],
                ["generated_at", datetime.now().isoformat(timespec="seconds")],
            ]

            return {
                "metadata": metadata,
                "header": ["column_index", "column_px", "column_unit", "bin_start", "bin_end", "count"],
                "rows": rows,
            }

        extra_tables = []
        if distribution_data.get("mode") == "Histogram":
            histogram_rows = []
            histogram_counts = distribution_data.get("histogram_counts")
            histogram_edges = distribution_data.get("histogram_edges")
            if histogram_counts is not None and histogram_edges is not None:
                for bin_index, count in enumerate(histogram_counts):
                    histogram_rows.append([
                        int(bin_index),
                        float(histogram_edges[bin_index]),
                        float(histogram_edges[bin_index + 1]),
                        int(count),
                    ])
            if histogram_rows:
                extra_tables.append(
                    {
                        "title": "Histogram Bin Counts",
                        "header": ["bin_index", "bin_start", "bin_end", "count"],
                        "rows": histogram_rows,
                    }
                )

        rows = []
        values = distribution_data["values"]
        sorted_values = distribution_data["sorted_values"]
        theoretical = distribution_data["theoretical_quantiles"]
        row_count = max(values.size, sorted_values.size)
        for idx in range(row_count):
            rows.append([
                idx,
                float(values[idx]) if idx < values.size else "",
                float(theoretical[idx]) if idx < theoretical.size else "",
            ])

        metadata = [
            ["export_type", "graph_distribution"],
            ["graph_view_mode", distribution_data["mode"]],
            ["distribution_kind", distribution_data["kind"]],
            ["sample_count", str(values.size)],
            ["mean", f"{distribution_data['mean']:.12g}"],
            ["std", f"{distribution_data['std']:.12g}"],
            ["skewness", f"{distribution_data['skewness']:.12g}"],
            ["excess_kurtosis", f"{distribution_data['excess_kurtosis']:.12g}"],
            ["generated_at", datetime.now().isoformat(timespec="seconds")],
        ]

        if distribution_data.get("combined_columns"):
            metadata.insert(3, ["combined_columns", "true"])
            metadata.insert(4, ["column_count", str(distribution_data["column_count"])])
            metadata.insert(5, ["valid_column_count", str(distribution_data["valid_column_count"])])
        else:
            metadata.insert(3, ["selected_column_px", f"{distribution_data['selected_x_px']:.12g}"])
            metadata.insert(4, ["selected_column_unit", f"{distribution_data['selected_x_units']:.12g}"])
            metadata.insert(5, ["selected_column_display", f"{distribution_data.get('selected_x_display', distribution_data['selected_x_px']):.12g}"])
            metadata.insert(6, ["column_input_mode", distribution_data.get("column_input_mode", "Pixel Values")])

        export_payload = {
            "metadata": metadata,
            "header": ["sample_index", "value", "theoretical_normal_quantile"],
            "rows": rows,
        }
        if extra_tables:
            export_payload["extra_tables"] = extra_tables
        return export_payload

    def _draw_centered_text_cv(self, image, text, center_x, baseline_y, font_scale=0.55, thickness=1, color=(0, 0, 0)):
        label = str(text)
        (text_w, _text_h), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, font_scale, thickness)
        text_x = int(round(center_x - (text_w / 2.0)))
        cv2.putText(image, label, (max(0, text_x), int(baseline_y)), cv2.FONT_HERSHEY_SIMPLEX, font_scale, color, thickness, cv2.LINE_AA)

    def _draw_vertical_text_cv(self, image, text, x_left, center_y, font_scale=0.55, thickness=1, color=(0, 0, 0)):
        label = str(text)
        (text_w, text_h), baseline = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, font_scale, thickness)
        temp_h = max(1, text_h + baseline + 8)
        temp_w = max(1, text_w + 8)
        temp = np.full((temp_h, temp_w, 3), 255, dtype=np.uint8)
        cv2.putText(temp, label, (2, text_h + 2), cv2.FONT_HERSHEY_SIMPLEX, font_scale, color, thickness, cv2.LINE_AA)
        rotated = cv2.rotate(temp, cv2.ROTATE_90_COUNTERCLOCKWISE)

        y_start = int(round(center_y - (rotated.shape[0] / 2.0)))
        x_start = int(x_left)
        y_end = y_start + rotated.shape[0]
        x_end = x_start + rotated.shape[1]

        clip_y0 = max(0, y_start)
        clip_x0 = max(0, x_start)
        clip_y1 = min(image.shape[0], y_end)
        clip_x1 = min(image.shape[1], x_end)
        if clip_y1 <= clip_y0 or clip_x1 <= clip_x0:
            return

        src_y0 = clip_y0 - y_start
        src_x0 = clip_x0 - x_start
        src_y1 = src_y0 + (clip_y1 - clip_y0)
        src_x1 = src_x0 + (clip_x1 - clip_x0)
        src = rotated[src_y0:src_y1, src_x0:src_x1]
        dst = image[clip_y0:clip_y1, clip_x0:clip_x1]
        mask = np.any(src < 250, axis=2)
        dst[mask] = src[mask]

    def _draw_right_aligned_text_cv(self, image, text, right_x, baseline_y, font_scale=0.45, thickness=1, color=(0, 0, 0)):
        label = str(text)
        (text_w, _text_h), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, font_scale, thickness)
        text_x = int(round(right_x - text_w))
        cv2.putText(image, label, (max(0, text_x), int(baseline_y)), cv2.FONT_HERSHEY_SIMPLEX, font_scale, color, thickness, cv2.LINE_AA)

    def render_graph_image(self, width=1600, height=1000):
        if str(self.app.graph_view_mode_var.get()).strip() != "Profile":
            return self.render_distribution_graph_image(width=width, height=height)

        app = self.app
        plot_data = self.build_plot_data()
        if plot_data is None:
            return None
        source = str(plot_data.get("source", "native")).strip()
        valid = plot_data["valid"]
        stdevs = plot_data["stdevs"]
        x_all_values = plot_data["x_all_values"]
        mean_values = plot_data["mean_values"]
        upper_values = plot_data["upper_values"]
        lower_values = plot_data["lower_values"]
        valid_idx = np.where(valid)[0]

        image = np.full((height, width, 3), 255, dtype=np.uint8)
        left, right, top, bottom = 140, 30, 50, 110
        plot_w = max(1, width - left - right)
        plot_h = max(1, height - top - bottom)

        y_range_values = np.concatenate((upper_values[valid], lower_values[valid]))
        limits = self.resolve_axis_limits(x_all_values[valid], y_range_values, y_pad=1.0)
        if limits is None:
            return None
        x_min, x_max, y_min, y_max = limits

        def x_to_px(x_value):
            if x_max <= x_min:
                return int(left + (plot_w / 2.0))
            return int(round(left + ((x_value - x_min) / (x_max - x_min)) * plot_w))

        def y_to_px(y_value):
            return int(round(top + ((y_max - y_value) / (y_max - y_min)) * plot_h))

        cv2.line(image, (left, top + plot_h), (left + plot_w, top + plot_h), (0, 0, 0), 1)
        cv2.line(image, (left, top), (left, top + plot_h), (0, 0, 0), 1)

        y_ticks = 5
        for i in range(y_ticks + 1):
            frac = i / y_ticks
            y_val = y_min + (y_max - y_min) * frac
            y_px = y_to_px(y_val)
            cv2.line(image, (left - 6, y_px), (left, y_px), (0, 0, 0), 1)
            label = self.format_graph_value(y_val)
            self._draw_right_aligned_text_cv(image, label, left - 10, y_px + 4)

        x_ticks = 6
        for i in range(x_ticks):
            frac = i / max(1, (x_ticks - 1))
            x_val = x_min + frac * (x_max - x_min)
            x_px = x_to_px(x_val)
            cv2.line(image, (x_px, top + plot_h), (x_px, top + plot_h + 6), (0, 0, 0), 1)
            cv2.putText(
                image,
                self.format_graph_value(x_val),
                (x_px - 12, top + plot_h + 24),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.45,
                (0, 0, 0),
                1,
                cv2.LINE_AA,
            )

        upper_pts = np.array([(x_to_px(x_all_values[idx]), y_to_px(upper_values[idx])) for idx in valid_idx], dtype=np.int32)
        lower_pts = np.array([(x_to_px(x_all_values[idx]), y_to_px(lower_values[idx])) for idx in valid_idx[::-1]], dtype=np.int32)
        band_poly = np.vstack((upper_pts, lower_pts))
        if band_poly.shape[0] >= 3:
            cv2.fillPoly(image, [band_poly], (255, 232, 207))

        mean_pts = np.array([(x_to_px(x_all_values[idx]), y_to_px(mean_values[idx])) for idx in valid_idx], dtype=np.int32)
        if mean_pts.shape[0] >= 2:
            cv2.polylines(image, [mean_pts], False, (189, 95, 0), 2)

        if app.show_best_fit_var.get():
            fit = self.compute_plot_fit(plot_data)
            if fit is not None:
                if source == "imported_csv":
                    fit_pts = np.array(
                        [
                            (x_to_px(x_all_values[idx]), y_to_px(float(fit["y_fit_values"][idx])))
                            for idx in valid_idx
                        ],
                        dtype=np.int32,
                    )
                else:
                    fit_pts = np.array(
                        [
                            (x_to_px(x_all_values[idx]), y_to_px(float(self.profile_y_value(fit["y_fit"][idx]))))
                            for idx in valid_idx
                        ],
                        dtype=np.int32,
                    )
                if fit_pts.shape[0] >= 2:
                    cv2.polylines(image, [fit_pts], False, (40, 40, 220), 2)
                cv2.putText(
                    image,
                    fit["equation"],
                    (left, top + 44),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.5,
                    (40, 40, 40),
                    1,
                    cv2.LINE_AA,
                )

        title = self.resolve_graph_title(plot_data.get("title", f"Final Mean Centerline with +/- {stdevs:g}sigma Band ({self.profile_axis_unit_label()})"))
        self._draw_centered_text_cv(image, title, width // 2, 30, font_scale=0.62, thickness=1)
        x_axis_text = app.graph_x_axis_label.get().strip() or self.profile_default_x_axis_label()
        y_axis_text = app.graph_y_axis_label.get().strip() or self.profile_default_y_axis_label()
        self._draw_centered_text_cv(image, x_axis_text, width // 2, height - 28, font_scale=0.55, thickness=1)
        self._draw_vertical_text_cv(image, y_axis_text, 8, top + (plot_h // 2), font_scale=0.55, thickness=1)
        return image

    def render_distribution_graph_image(self, width=1600, height=1000):
        app = self.app
        mode = str(app.graph_view_mode_var.get()).strip()
        distribution_data = self.build_distribution_data()
        if distribution_data is None:
            return None

        image = np.full((height, width, 3), 255, dtype=np.uint8)
        left, right, top, bottom = 140, 30, 50, 110
        plot_w = max(1, width - left - right)
        plot_h = max(1, height - top - bottom)

        if mode == "Histogram":
            if distribution_data.get("all_columns"):
                edges = distribution_data["histogram_edges"]
                counts_matrix = distribution_data["histogram_matrix"]
                x_column_display = distribution_data["x_column_display"]
                max_count = int(np.max(counts_matrix)) if counts_matrix.size > 0 else 0
                if max_count <= 0:
                    return None

                if x_column_display.size == 1:
                    x_edges = np.array([x_column_display[0] - 0.5, x_column_display[0] + 0.5], dtype=np.float64)
                else:
                    midpoints = 0.5 * (x_column_display[:-1] + x_column_display[1:])
                    first_edge = x_column_display[0] - (midpoints[0] - x_column_display[0])
                    last_edge = x_column_display[-1] + (x_column_display[-1] - midpoints[-1])
                    x_edges = np.concatenate(([first_edge], midpoints, [last_edge]))

                x_min = float(x_edges[0])
                x_max = float(x_edges[-1])
                y_min = float(edges[0])
                y_max = float(edges[-1])

                def x_to_px(x_value):
                    if x_max <= x_min:
                        return int(left + (plot_w / 2.0))
                    return int(round(left + ((x_value - x_min) / (x_max - x_min)) * plot_w))

                def y_to_px(y_value):
                    return int(round(top + ((y_max - y_value) / max(y_max - y_min, 1.0)) * plot_h))

                cv2.line(image, (left, top + plot_h), (left + plot_w, top + plot_h), (0, 0, 0), 1)
                cv2.line(image, (left, top), (left, top + plot_h), (0, 0, 0), 1)

                for col_idx in range(counts_matrix.shape[0]):
                    x0 = x_to_px(float(x_edges[col_idx]))
                    x1 = x_to_px(float(x_edges[col_idx + 1]))
                    for bin_idx in range(counts_matrix.shape[1]):
                        count = int(counts_matrix[col_idx, bin_idx])
                        if count <= 0:
                            continue
                        y0 = y_to_px(float(edges[bin_idx + 1]))
                        y1 = y_to_px(float(edges[bin_idx]))
                        norm = max(0.0, min(1.0, count / max_count))
                        color = (
                            int(round(255 - 21 * norm)),
                            int(round(251 - 196 * norm)),
                            int(round(248 - 184 * norm)),
                        )
                        cv2.rectangle(image, (x0, y0), (x1, y1), color, -1)

                y_ticks = 5
                for i in range(y_ticks + 1):
                    frac = i / y_ticks
                    y_val = y_min + (y_max - y_min) * frac
                    y_px = y_to_px(y_val)
                    cv2.line(image, (left - 6, y_px), (left, y_px), (0, 0, 0), 1)
                    self._draw_right_aligned_text_cv(image, self.format_graph_value(y_val), left - 10, y_px + 4)

                x_ticks = 6
                for i in range(x_ticks):
                    frac = i / max(1, (x_ticks - 1))
                    x_val = x_min + frac * (x_max - x_min)
                    x_px = x_to_px(x_val)
                    cv2.line(image, (x_px, top + plot_h), (x_px, top + plot_h + 6), (0, 0, 0), 1)
                    cv2.putText(image, self.format_graph_value(x_val), (x_px - 12, top + plot_h + 24), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 0, 0), 1, cv2.LINE_AA)

                title = (
                    f"Histogram Heatmap of {distribution_data['kind']} Across Columns "
                    f"({distribution_data['sample_count']} samples)"
                )
                x_axis_text = distribution_data["default_x_label"]
                y_axis_text = distribution_data["default_y_label"]
            else:
                edges = distribution_data["histogram_edges"]
                counts = distribution_data["histogram_counts"]
                limits = self.resolve_axis_limits(edges, np.append(counts, 0), y_pad=1.0)
                if limits is None:
                    return None
                x_min, x_max, _y_min, y_max = limits
                y_min = 0.0

                def x_to_px(x_value):
                    if x_max <= x_min:
                        return int(left + (plot_w / 2.0))
                    return int(round(left + ((x_value - x_min) / (x_max - x_min)) * plot_w))

                def y_to_px(y_value):
                    return int(round(top + ((y_max - y_value) / max(y_max - y_min, 1.0)) * plot_h))

                cv2.line(image, (left, top + plot_h), (left + plot_w, top + plot_h), (0, 0, 0), 1)
                cv2.line(image, (left, top), (left, top + plot_h), (0, 0, 0), 1)

                for idx, count in enumerate(counts):
                    x0 = x_to_px(edges[idx])
                    x1 = x_to_px(edges[idx + 1])
                    y1 = y_to_px(float(count))
                    cv2.rectangle(image, (x0, y1), (x1, top + plot_h), (189, 126, 74), 1)
                    cv2.rectangle(image, (x0, y1), (x1, top + plot_h), (255, 232, 207), -1)
                    cv2.rectangle(image, (x0, y1), (x1, top + plot_h), (189, 126, 74), 1)

                y_ticks = 5
                for i in range(y_ticks + 1):
                    frac = i / y_ticks
                    y_val = y_min + (y_max - y_min) * frac
                    y_px = y_to_px(y_val)
                    cv2.line(image, (left - 6, y_px), (left, y_px), (0, 0, 0), 1)
                    self._draw_right_aligned_text_cv(image, self.format_graph_value(y_val), left - 10, y_px + 4)

                x_ticks = 6
                for i in range(x_ticks):
                    frac = i / max(1, (x_ticks - 1))
                    x_val = x_min + frac * (x_max - x_min)
                    x_px = x_to_px(x_val)
                    cv2.line(image, (x_px, top + plot_h), (x_px, top + plot_h + 6), (0, 0, 0), 1)
                    cv2.putText(image, self.format_graph_value(x_val), (x_px - 12, top + plot_h + 24), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 0, 0), 1, cv2.LINE_AA)

                std_value = distribution_data["std"]
                if std_value > 0:
                    x_curve = np.linspace(x_min, x_max, 200)
                    bin_width = edges[1] - edges[0] if len(edges) > 1 else 1.0
                    y_curve = np.array([
                        distribution_data["values"].size * bin_width * NormalDist(mu=distribution_data["mean"], sigma=std_value).pdf(float(x_val))
                        for x_val in x_curve
                    ], dtype=np.float64)
                    points = np.array([(x_to_px(x_val), y_to_px(y_val)) for x_val, y_val in zip(x_curve, y_curve)], dtype=np.int32)
                    if points.shape[0] >= 2:
                        cv2.polylines(image, [points], False, (40, 40, 220), 2)

                if distribution_data.get("combined_columns"):
                    title = (
                        f"Histogram of {distribution_data['kind']} Across All Columns "
                        f"({distribution_data['values'].size} samples)"
                    )
                else:
                    title = (
                        f"Histogram of {distribution_data['kind']} at x={distribution_data['selected_x_display']:.3g} "
                        f"{distribution_data.get('column_input_suffix', 'px')} "
                        f"({distribution_data['values'].size} samples)"
                    )
                x_axis_text = distribution_data["default_x_label"]
                y_axis_text = distribution_data["default_y_label"]
        else:
            theoretical = distribution_data["theoretical_quantiles"]
            observed = distribution_data["sorted_values"]
            combined = np.concatenate((theoretical, observed))
            limits = self.resolve_axis_limits(combined, combined, y_pad=1.0)
            if limits is None:
                return None
            x_min, x_max, y_min, y_max = limits

            def x_to_px(x_value):
                if x_max <= x_min:
                    return int(left + (plot_w / 2.0))
                return int(round(left + ((x_value - x_min) / (x_max - x_min)) * plot_w))

            def y_to_px(y_value):
                return int(round(top + ((y_max - y_value) / max(y_max - y_min, 1.0)) * plot_h))

            cv2.line(image, (left, top + plot_h), (left + plot_w, top + plot_h), (0, 0, 0), 1)
            cv2.line(image, (left, top), (left, top + plot_h), (0, 0, 0), 1)

            diag_start_x = max(x_min, y_min)
            diag_end_x = min(x_max, y_max)
            if diag_end_x > diag_start_x:
                cv2.line(image, (x_to_px(diag_start_x), y_to_px(diag_start_x)), (x_to_px(diag_end_x), y_to_px(diag_end_x)), (140, 140, 140), 1)

            points = np.array([(x_to_px(x_val), y_to_px(y_val)) for x_val, y_val in zip(theoretical, observed)], dtype=np.int32)
            for x_px, y_px in points:
                cv2.circle(image, (int(x_px), int(y_px)), 3, (189, 95, 0), -1)

            y_ticks = 5
            for i in range(y_ticks + 1):
                frac = i / y_ticks
                y_val = y_min + (y_max - y_min) * frac
                y_px = y_to_px(y_val)
                cv2.line(image, (left - 6, y_px), (left, y_px), (0, 0, 0), 1)
                self._draw_right_aligned_text_cv(image, self.format_graph_value(y_val), left - 10, y_px + 4)

            x_ticks = 6
            for i in range(x_ticks):
                frac = i / max(1, (x_ticks - 1))
                x_val = x_min + frac * (x_max - x_min)
                x_px = x_to_px(x_val)
                cv2.line(image, (x_px, top + plot_h), (x_px, top + plot_h + 6), (0, 0, 0), 1)
                cv2.putText(image, self.format_graph_value(x_val), (x_px - 12, top + plot_h + 24), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 0, 0), 1, cv2.LINE_AA)

            if distribution_data.get("combined_columns"):
                title = f"Normal Q-Q Plot of {distribution_data['kind']} Across All Columns"
            else:
                title = (
                    f"Normal Q-Q Plot of {distribution_data['kind']} at x={distribution_data['selected_x_display']:.3g} "
                    f"{distribution_data.get('column_input_suffix', 'px')}"
                )
            x_axis_text = app.graph_x_axis_label.get().strip() or "Theoretical Normal Quantile"
            y_axis_text = app.graph_y_axis_label.get().strip() or "Observed Quantile"

        self._draw_centered_text_cv(image, self.resolve_graph_title(title), width // 2, 30, font_scale=0.62, thickness=1)
        self._draw_centered_text_cv(image, x_axis_text, width // 2, height - 28, font_scale=0.55, thickness=1)
        self._draw_vertical_text_cv(image, y_axis_text, 8, top + (plot_h // 2), font_scale=0.55, thickness=1)
        return image

    def save_graph_image(self):
        app = self.app
        default_dir = app.output_dir.get().strip() or os.path.dirname(__file__)
        default_file = f"{app.output_name_entry.get().strip() or 'analysis_output'}_graph.png"
        file_path = filedialog.asksaveasfilename(
            title="Save graph image",
            initialdir=default_dir,
            initialfile=default_file,
            defaultextension=".png",
            filetypes=[
                ("PNG image", "*.png"),
                ("JPEG image", "*.jpg *.jpeg"),
                ("Bitmap image", "*.bmp"),
                ("TIFF image", "*.tif *.tiff"),
            ],
        )
        if not file_path:
            return
        try:
            self.save_graph_image_to_path(file_path)
            messagebox.showinfo("Graph saved", f"Saved graph image to:\n{file_path}")
        except (OSError, ValueError) as exc:
            messagebox.showerror("Save failed", str(exc))

    def save_graph_image_to_path(self, file_path):
        image = self.render_graph_image()
        if image is None:
            raise ValueError("Run analysis or import a CSV first to generate graph data.")

        output_dir = os.path.dirname(file_path) or "."
        os.makedirs(output_dir, exist_ok=True)
        if not cv2.imwrite(file_path, image):
            raise OSError(f"Could not save graph image:\n{file_path}")
        return file_path

    def _write_graph_data_csv(self, file_path, export_data):
        with open(file_path, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            metadata = export_data.get("metadata", [])
            data_rows = export_data.get("rows", [])

            if "pixel_header" in export_data and "unit_header" in export_data:
                pixel_header = export_data["pixel_header"]
                unit_header = export_data["unit_header"]
                _raw = getattr(self.app, "final_centerline_samples", None)
                histogram_export = (
                    self._build_combined_distribution_export(_raw)
                    if _raw is not None
                    else None
                )
                histogram_header = []
                histogram_rows = []
                histogram_metadata = []
                if histogram_export is not None:
                    histogram_header = list(histogram_export.get("header", []))
                    histogram_rows = list(histogram_export.get("rows", []))
                    histogram_metadata = list(histogram_export.get("metadata", []))

                if histogram_metadata:
                    metadata = list(metadata) + [
                        [f"histogram_{str(key)}", str(value)] for key, value in histogram_metadata
                    ]

                unit_label_text = (self.app.graph_unit_label or "Actual Units").strip()
                position_title = ["Position Data (Pixels)"] + [""] * (len(pixel_header) - 1)
                unit_title = [f"Position Data ({unit_label_text})"] + [""] * (len(unit_header) - 1)
                histogram_title = ["Histogram Data"] + [""] * (len(histogram_header) - 1) if histogram_header else []

                title_row = ["Metadata", "Value", "", *position_title, "", *unit_title]
                if histogram_header:
                    title_row.extend(["", *histogram_title])
                writer.writerow(title_row)

                header_row = ["metadata_key", "metadata_value", "", *pixel_header, "", *unit_header]
                if histogram_header:
                    header_row.extend(["", *histogram_header])
                writer.writerow(header_row)

                row_count = max(len(data_rows), len(metadata), len(histogram_rows))
                for row_idx in range(row_count):
                    if row_idx < len(metadata):
                        metadata_key, metadata_value = metadata[row_idx]
                    else:
                        metadata_key, metadata_value = "", ""

                    if row_idx < len(data_rows):
                        pixel_part = data_rows[row_idx].get("pixel", [""] * len(pixel_header))
                        unit_part = data_rows[row_idx].get("unit", [""] * len(unit_header))
                    else:
                        pixel_part = [""] * len(pixel_header)
                        unit_part = [""] * len(unit_header)

                    row = [metadata_key, metadata_value, "", *pixel_part, "", *unit_part]
                    if histogram_header:
                        histogram_part = histogram_rows[row_idx] if row_idx < len(histogram_rows) else [""] * len(histogram_header)
                        row.extend(["", *histogram_part])
                    writer.writerow(row)
            else:
                combined_header = ["metadata_key", "metadata_value", *export_data["header"]]
                writer.writerow(combined_header)

                row_count = max(len(data_rows), len(metadata))
                for row_index in range(row_count):
                    data_part = data_rows[row_index] if row_index < len(data_rows) else [""] * len(export_data["header"])
                    if row_index < len(metadata):
                        metadata_key, metadata_value = metadata[row_index]
                    else:
                        metadata_key, metadata_value = "", ""
                    writer.writerow([metadata_key, metadata_value, *data_part])

                # Preserve the existing CSV table format and append optional
                # supplemental sections for view-specific data (for example, histogram bins).
                extra_tables = export_data.get("extra_tables", [])
                for table in extra_tables:
                    table_header = table.get("header", [])
                    table_rows = table.get("rows", [])
                    if not table_header:
                        continue
                    writer.writerow([])
                    table_title = str(table.get("title", "Additional Data")).strip()
                    writer.writerow([table_title])
                    writer.writerow(table_header)
                    for table_row in table_rows:
                        writer.writerow(table_row)

    def save_graph_data_csv_to_path(self, file_path):
        export_data = self.build_graph_export_rows()
        if export_data is None:
            raise ValueError("Run analysis or import a CSV first to generate graph data.")

        output_dir = os.path.dirname(file_path) or "."
        os.makedirs(output_dir, exist_ok=True)
        self._write_graph_data_csv(file_path, export_data)
        return file_path

    def save_graph_data_csv(self):
        app = self.app
        default_dir = app.output_dir.get().strip() or os.path.dirname(__file__)
        default_file = f"{app.output_name_entry.get().strip() or 'analysis_output'}_graph_data.csv"
        file_path = filedialog.asksaveasfilename(
            title="Save graph data (CSV)",
            initialdir=default_dir,
            initialfile=default_file,
            defaultextension=".csv",
            filetypes=[("CSV file", "*.csv")],
        )
        if not file_path:
            return

        try:
            self.save_graph_data_csv_to_path(file_path)
            messagebox.showinfo("Graph data saved", f"Saved graph CSV to:\n{file_path}")
        except (OSError, ValueError) as exc:
            messagebox.showerror("Save failed", str(exc))
